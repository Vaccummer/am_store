import pandas as pd
from glob import glob
import webbrowser
import openpyxl
import subprocess
import os
import copy
from typing import Literal, Optional, Tuple, Union, List, OrderedDict
import re
from PySide2.QtGui import *
from PySide2.QtWidgets import *
from PySide2.QtGui import *
from PySide2.QtCore import *
import warnings
import math
from sympy import symbols, sympify
from am_store.common_tools import yml, AMPATH
import paramiko
import stat
import ctypes
import asyncssh
import aiofiles
import asyncio
import pathlib
from pathlib import Path
import time
from abc import ABC, abstractmethod

def is_path(string_f, exist_check:bool=False):
    # To judge whether a variable is Path or not
    if isinstance(string_f, pathlib.Path):
        if exist_check:
            return string_f.exists()
        else:
            return True
    if not isinstance(string_f, str):
        return False
    if not exist_check:
        sign_f = False
        if os.path.isabs(string_f):
            sign_f = True
        if sign_f:
            return True
        pattern_f = r'^[~./\\]|^[a-zA-Z]:\\+'
        return bool(re.match(pattern_f, string_f))
    else:
        return os.path.exists(string_f)

def path_format(string_f:str):
    # turn a string into a formal path of OS
    if not isinstance(string_f, str):
        raise ValueError(f"Input of path_format function should be a string, but get {type(string_f)}")
    else:
        sep = os.sep
        string_f = rf"{string_f}"
        path_t1 = string_f.replace('/', sep)
        path_t2 = path_t1.replace('\\', sep)
        return path_t2
def path_join(*args, mkdir:Literal[True, False]=True):
    # function to join str to path
    if not all(isinstance(arg, str) for arg in args):
        raise TypeError("All arguments must be of type str")
    path = os.path.join(*args)
    check_flag_f = 0
    
    pattern_f = r'^.+\.\d*[a-zA-Z]+\w*$'
    base_f = os.path.basename(path)
    
    if re.match(pattern_f, base_f):
        check_flag_f = 1
    else:
        check_flag_f = 0
    
    if mkdir:
        if check_flag_f == 0:
            os.makedirs(path, exist_ok=True)
        else:
            path_up = os.path.dirname(path)
            os.makedirs(path_up, exist_ok=True)
    return path
def web_open(url_f):
    # web lauch
    #subprocess.run(["start", r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe", url_f], shell=True)
    webbrowser.open(url_f)
def list_overlap(list_list_f):
    # To remove overlap of multi calsses
    return list(set.intersection(*map(set, list_list_f)))
def opto_excel_process(excel_path_f, data_f=None):
    # Read Optogenetic Setting Excel
    if not data_f:
        wb = openpyxl.load_workbook(excel_path_f, data_only=True)
        ws = wb.worksheets[0]
        empty_count_f = 0
        para_dict_f = {}
        for i_f in range(1, 10000):
            para_n = ws.cell(row=i_f, column=1).value
            para_v = ws.cell(row=i_f, column=2).value
            if empty_count_f >= 10:
                break
            if para_n and para_v:
                para_dict_f[para_n] = para_v
            else:
                empty_count_f += 1
        return para_dict_f
    else:
        wb = openpyxl.load_workbook(excel_path_f, data_only=True)
        ws = wb.worksheets[0]
        para_dict_old = {}
        empty_count_f = 0
        for i_f in range(1, 10000):
            para_n = ws.cell(row=i_f, column=1).value
            para_v = ws.cell(row=i_f, column=2).value
            if empty_count_f >= 10:
                break
            if para_n and para_v:
                para_dict_old[para_n] = i_f
            else:
                empty_count_f += 1
        for item_f, value_f in data_f.items():
            if item_f in para_dict_old.keys():
                ws.cell(row=para_dict_old[item_f], column=2, value=value_f)
        wb.save(excel_path_f)
def command_conduct(prompt_f):
    # Conduct EXE file or Open doc in Explorer
    if is_url(prompt_f):
        chrome_cmd = "start msedge --new-window " + prompt_f
        subprocess.Popen(chrome_cmd, shell=True)
    elif os.path.exists(prompt_f):
        subprocess.Popen(["explorer", prompt_f])
def rm_dp_list_elem(list_f, reverse=True):
    # To remove overlap of multi calsses, remain elements which rank upper
    a_no_duplicates = []
    if reverse:
        for i in reversed(list_f):
            if i not in a_no_duplicates:
                a_no_duplicates.insert(0, i)
    else:
        for i in list_f:
            if i not in a_no_duplicates:
                a_no_duplicates.append(i)
    return a_no_duplicates
def dir_name_sort(dir_list_f, prompt_f):
    # Sorted Association Words according to the location of prompt
    order_list = []
    for dir_f in dir_list_f:
        try:
            order_list.append([dir_f.index(prompt_f), dir_f])
        except Exception:
            order_list.append([999999, dir_f])
    return [x[1] for x in sorted(order_list, key=lambda x:x[0])]
def resize(data_f, para_dict_f, type_f):
    # resize geometry
    data_f = copy.deepcopy(data_f)
    scrx = para_dict_f['scr_x']
    scry = para_dict_f['scr_y']
    res_x = para_dict_f['res_x']
    res_y = para_dict_f['res_y']
    gap = para_dict_f['gap']
    het = para_dict_f['het']
    radius = para_dict_f['srh_r']
    win_x = para_dict_f['win_x']
    win_y = para_dict_f['win_y']
    if type_f == 'font':
        for value_if in data_f.values():
            temp_num = value_if['PixelSize']
            num_d = int(temp_num*min(res_x, res_y))
            value_if['PixelSize'] = num_d
            
    elif type_f == 'window':
        for dict_if in data_f.values():
            for key_f, value_f in dict_if.items():
                if not value_f:
                    continue
                else:
                    x, y, w, h = value_f
                    dict_t = dict()
                    exec(f'x_t=int({x})\ny_t=int({y})\nw_t=int({w})\nh_t=int({h})', locals(), dict_t)
                    dict_if[key_f] = [dict_t['x_t'], dict_t['y_t'], dict_t['w_t'], dict_t['h_t']]
    return data_f
def config_process(config_f:dict):
    ## process config file, resize and cal the geometry of windows
    # HyperSizeSet Edit
    hypersizeset = config_f['HyperSizeSet']
    scr_x, scr_y = get_screen_size()
    res_x = scr_x/2560
    res_y = scr_y/1440
    win_x = hypersizeset['win_x']
    win_y = hypersizeset['win_y']
    ## resize
    hypersizeset['gap'] = int(res_x*hypersizeset['gap'])
    hypersizeset['het'] = int(res_x*hypersizeset['het'])
    hypersizeset['srh_r'] = int(hypersizeset['srh_r']*(res_x+res_y)/2)

    gap = hypersizeset['gap']
    het = hypersizeset['het']
    srh_r = hypersizeset['srh_r']

    para_wait_for_edit = ['scr_x', 'scr_y', 'res_x', 'res_y']
    for para_if in para_wait_for_edit:
        hypersizeset[para_if] = hypersizeset[para_if] if hypersizeset[para_if] else locals()[para_if]

    # LauncherSize Edit
    hyperpara_dict = {"scr_x":scr_x, "scr_y":scr_y, "res_x":res_x, "res_y":res_y, 
                      "gap":gap, "het":het, "srh_r":srh_r, "win_x":win_x, "win_y":win_y}
    
    launchersize = config_f['LauncherSize']
    gb = {"__builtins__": None}
    lc = hyperpara_dict
    #eval_f = partial(eval, globals={"__builtins__": None}, locals=hyperpara_dict)
    for mode_name, window_geometry in launchersize.items():
        for window_name, geometry in window_geometry.items():
            if not geometry:
                continue
            x, y, w, h = str(geometry[0]), str(geometry[1]), str(geometry[2]), str(geometry[3])
            x_t, y_t, w_t, h_t = eval(x,gb,lc), eval(y,gb,lc), eval(w,gb,lc), eval(h,gb,lc)
            launchersize[mode_name][window_name] = [x_t, y_t, w_t, h_t]
    
    ## path translate
    launcher_path_set = config_f['LauncherPath']
    launcher_wkdir = path_join(config_f['Common']['work_dir'], "load")
    launcher_path_set = path_translate(launcher_path_set, launcher_wkdir)
    config_f['LauncherPath'] = launcher_path_set

    media_path_set = config_f['MediaPath']
    media_wkdir = path_join(config_f['Common']['work_dir'], "load", "player")
    media_path_set = path_translate(media_path_set, media_wkdir)
    config_f['MediaPath'] = media_path_set
    return config_f
def path_translate(path_dict_f, work_dir_f):
    path_n = dicta.flatten_dict(copy.deepcopy(path_dict_f))
    for key_f, value_f in path_n.items():
        if not value_f:
            path_l = glob(path_join(work_dir_f,*key_f[:-1], f"{key_f[-1]}.*", mkdir=False))
            path_n[key_f] = path_l[0] if path_l else None
        else:
            path_l = glob(path_join(work_dir_f, *key_f[:-1], value_f, "*", mkdir=False))
            path_n[key_f] = path_l if path_l else []
    return dicta.unflatten_dict(path_n)
def path_trans(path_dict_f, work_dir_f):
    # search for target path
    for window_ni, dict_if in path_dict_f.items():
        for key_if, value_if in dict_if.items():
            if not value_if:
                path_t = glob(os.path.join(work_dir_f, window_ni, f"{key_if}.*"))
                if len(path_t) == 1:
                    path_dict_f[window_ni][key_if] = path_t[0]
                else:
                    print(f"{window_ni}-{key_if}: No Target Detected!")
            else:
                path_t = glob(os.path.join(work_dir_f, window_ni, value_if, "*"))
                if path_t:
                    path_dict_f[window_ni][key_if] = path_t
                else:
                    print(f"{window_ni}-{value_if}: No Target Detected!")
    return path_dict_f
def nan_to_sign(list_f):
    # repalce nan with None
    return ['' if pd.isna(i) else str(i) for i in list_f]
def is_url(str_f:str):
    # judge a variable whether is a url or not
    if not isinstance(str_f, str):
        return False
    url_pattern = re.compile(
    r'^(http|https)://'  # 匹配协议部分 http:// 或 https://
    r'([a-zA-Z0-9.-]+)'  # 匹配域名部分
    r'(\.[a-zA-Z]{2,})'   # 匹配顶级域名部分
    )
    match_f = url_pattern.match(str_f)
    return bool(match_f)
def get_screen_size2(target: Literal['pixel', 'physical', "dpi"]="pixel"):
    app = QApplication.instance() if QApplication.instance() else QApplication([])
    allowed_values = ["pixel", "physical","dpi"]
    assert target in allowed_values, f"Invalid value: {target}. Allowed values are: {allowed_values}"
    screen = app.primaryScreen()  # 获取主屏幕对象
    match target:
        case "pixel":
            pixel_size = screen.size()
            width_px = pixel_size.width()
            height_px = pixel_size.height()
            app.quit()
            return (width_px, height_px)
        case "physical":
            physical_size = screen.physicalSize()
            width_mm = physical_size.width()
            height_mm = physical_size.height()
            app.quit()
            return [width_mm, height_mm]
        case "dpi":
            app.quit()
            return screen.physicalDotsPerInch()
def get_screen_size(target: Literal['pixel', 'physical', "dpi"]="pixel"):
    user32 = ctypes.windll.user32
    user32.SetProcessDPIAware()  # 使程序支持高 DPI
    width = user32.GetSystemMetrics(0)
    height = user32.GetSystemMetrics(1)
    return width, height
def excel_to_df(excel_path_f:str, region:str, sheet_name_f:Union[int, str, list]='Sheet1'):  
    # Read XLSX doc and transform it into dataframe
    # region format : 'A1:Z7'
    # sheet_name_f : Sheet name or Sheet order number(start from 1)
    import pandas as pd
    assert is_path(excel_path_f, exist_check=True), f"Invalid path: {excel_path_f}. PATH NOT EXISTS!"
    try:
        dataframe_f = []
        if sheet_name_f == "all":
            workbook = openpyxl.load_workbook(excel_path_f)
            sheet_name_f = workbook.sheetnames
        if isinstance(sheet_name_f, list):
            for sheet_i in sheet_name_f:
                dataframe_i = pd.read_excel(excel_path_f, sheet_name=sheet_i, usecols=region)
                dataframe_f.append(dataframe_i)
            dataframe_f = pd.concat(dataframe_f, ignore_index=True)
        else:
            dataframe_f = pd.read_excel(excel_path_f, sheet_name=sheet_name_f, usecols=region)
    except Exception as e:
        dataframe_f = None
        warnings.warn(f"excel_to_df('{excel_path_f}') encounter error {e}, return None in default")
    return dataframe_f

def font_get(para_dict_f={"Family":None,
                        'PointSize':None,
                        'PixelSize':None,
                        'Bold':None,
                        'Italic':None,
                        'Weight':None,
                        'Underline':None,
                        'LetterSpacing':None,
                        'WordSpacing':None,
                        'Stretch':None,
                        'StrikeOut':None}):
    # produce font with family and size
    if not para_dict_f:
        return QFont()
    if isinstance(para_dict_f, list):
        para_dict_f = {i[0]:i[1] for i in para_dict_f}
    font_f = QFont()
    font_f.setStyleStrategy(QFont.PreferAntialias)
    allowed_keys = ["Family", "PointSize", "Bold", "Italic", "Weight", "Underline", "StrikeOut", 
                    "LetterSpacing", "WordSpacing", "Stretch", 'PixelSize']
    for key, value in para_dict_f.items():
        if (key in allowed_keys) and value:
            dict_f = {'font_f':font_f,
                      'value':value}
            exec(f'font_f.set{key}(value)', dict_f)
    
    return font_f

def parse_ssh_config(file_path, fliter:Literal['all'] | None | List[str]):
    hosts = {}
    try:
        with open(file_path, 'r') as file:
            current_host = None
            host_config = {}
            for line in file:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if line.lower().startswith('host '):
                    if current_host:
                        hosts[current_host] = host_config
                    current_host = line.split()[1]
                    host_config = {}
                    continue
                match = re.match(r'(\w+)\s+(.+)', line)
                if match:
                    key, value = match.groups()
                    host_config[key] = value
            if current_host:
                hosts[current_host] = host_config
        if fliter == 'all':
            return hosts
        elif fliter is None:
            return {}
        else:
            dict_t = {}
            for name_i in fliter:
                host_i = hosts.get(name_i, None)
                if host_i is not None:
                    dict_t[name_i] = host_i
            return dict_t
    except Exception as e:
        return {}

def extract_icon(exe_icon_getter, exe_path, target_icon_path):
    commands = [exe_icon_getter, exe_path, target_icon_path]
    try:
        result = subprocess.check_output(commands, cwd=os.path.dirname(exe_icon_getter)).decode('gbk')
        return result and "图标已保存为" in result
    except subprocess.CalledProcessError:
        return False
        
def amlayoutH(align_h:Literal['l','r','c']=None, align_v:Literal['t','b','c']=None, spacing:int=None)->QHBoxLayout:
    lo = QHBoxLayout()
    match align_h:
        case 'l':
            lo.setAlignment(Qt.AlignLeft)
        case "c":
            lo.setAlignment(Qt.AlignHCenter)
        case "r":
            lo.setAlignment(Qt.AlignRight)
        case _:
            pass
    match align_v:
        case "t":
            lo.setAlignment(Qt.AlignTop)
        case "b":
            lo.setAlignment(Qt.AlignBottom)
        case "c":
            lo.setAlignment(Qt.AlignVCenter)
    if spacing:
        lo.setSpacing(int(spacing))
    return lo
def amlayoutV(align_h:Literal['l','r','c']=None, align_v:Literal['t','b','c']=None, spacing:int=None)->QVBoxLayout:
    lo = QVBoxLayout()
    match align_h:
        case 'l':
            lo.setAlignment(Qt.AlignLeft)
        case "c":
            lo.setAlignment(Qt.AlignHCenter)
        case "r":
            lo.setAlignment(Qt.AlignRight)
        case _:
            pass
    match align_v:
        case "t":
            lo.setAlignment(Qt.AlignTop)
        case "b":
            lo.setAlignment(Qt.AlignBottom)
        case "c":
            lo.setAlignment(Qt.AlignVCenter)
    if spacing:
        lo.setSpacing(int(spacing))
    return lo

def add_obj(*args, parent_f:Union[QHBoxLayout, QVBoxLayout]):
    for arg_i in args:
        if isinstance(arg_i, QHBoxLayout):
            parent_f.addLayout(arg_i)
        elif isinstance(arg_i, QVBoxLayout):
            parent_f.addLayout(arg_i)
        else:
            parent_f.addWidget(arg_i)
    return parent_f

def setStyle(widget:QWidget, style_string:str, *args):
    """ widget is the target widget, style_string is the style string with format {}
    num of {} in style_string should be equal to args num
    **use partial to fix the widget and style_string
    """
    widget.setStyleSheet(style_string.format(*args))

def enlarge_list(value_f, length_f:int):
    if not isinstance(value_f, list):
        return [value_f]*length_f
    else:
        return value_f + [value_f[-1]]*(length_f-len(value_f))

def style_make(config:dict):
    """
    config must be a dict, key is obj name, value is a dict with key-value pairs
    """
    style = ""
    for obj_name, style_dict in config.items():
        style += f"{obj_name} {{\n"
        for key, value in style_dict.items():
            style += f"{key}: {value};\n"
        style += "}"
    return style

class atuple(tuple):
    def __new__(cls, *args):
        if len(args) == 1 and isinstance(args[0], list):
            return super().__new__(cls, *args)
        else:
            return super().__new__(cls, args)
    
    def __or__(self, other):
        if isinstance(other, atuple):
            return atuple(list(self) + list(other))
        else:
            return Atuple(list(self) + [other])

    def __hash__(self):
        return hash(tuple(Atuple('Atuple') | self))

class alist(list):
    def __init__(cls, *args):
        if len(args) == 1:
            return super().__init__(*args)
        else:
            return super().__init__(args)
    
    def __eq__(self, value):
        if not isinstance(value, list):
            return None
        elif len(self) != len(value):
            return None
        else:
            for i in range(len(self)):
                if self[i] != value[i]:
                    return False
            return True

class adict(dict):
    def __init__(self, dict_f:dict):
        self.dict_t = dict_f
    
    def __getitem__(self, key):
        if not isinstance(key, atuple):
            return self.dict_t.get(key, None)
        else:
            dict_n = self.dict_t
            for key_i in key:
                dict_n = dict_n.get(key_i, None)
                if dict_n is None:
                    return None
            return dict_n