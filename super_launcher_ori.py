from PySide2.QtWidgets import QApplication, QVBoxLayout, QLineEdit, QListWidget, QListWidgetItem, QApplication, QMainWindow, QComboBox, QVBoxLayout, QLabel
from PySide2.QtWidgets import QWidget, QTextEdit, QProgressBar, QToolButton, QPushButton, QFrame, QSystemTrayIcon, QMenu
from PySide2.QtWidgets import QVBoxLayout, QHBoxLayout, QFileDialog, QGraphicsBlurEffect, QScrollArea, QAction, QShortcut
from PySide2.QtCore import Qt, QEvent, QObject, Signal, QSize, Slot, QThread
from PySide2.QtCore import QTimer, QPropertyAnimation, QEasingCurve, QSequentialAnimationGroup
from PySide2.QtGui import QPixmap, QPalette, QIcon, QFont, QScreen, QFontMetrics, QWindow, QKeySequence
from PySide2.QtGui import QPainter, QColor, QBrush, QPen, QLinearGradient, QTextCharFormat, QTextCursor
import subprocess
import os
import multiprocessing
import time
import pandas as pd
import sys
import webbrowser
import re
import shutil
import random
import threading
from queue import Queue
from glob import glob
import openpyxl
import copy
import math
from functools import partial
from Data_Manage_System import opto_main
from huggingface_gpt import HuggingFaceGPT
from music_player_ui_formal import MusicPlayer
from Ameng import path_format, is_url, is_path, restart_program, excel_to_df, get_screen_size, font_get
import webbrowser
import atexit


# web lauch
def web_open(url_f):
    #subprocess.run(["start", r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe", url_f], shell=True)
    webbrowser.open(url_f)
# To remove overlap of multi calsses
def list_overlap(list_list_f):
    return list(set.intersection(*map(set, list_list_f)))
# Read Optogenetic Setting Excel
def opto_excel_process(excel_path_f, data_f=None):
    if not data_f:
        wb = openpyxl.load_workbook(excel_path_f, data_only=True)
        ws = wb.worksheets[0]
        empty_count_f = 0
        para_dict_f = {}
        for i_f in range(1, 10000):
            para_n = ws.cell(row=i_f, column=1).value
            para_v = ws.cell(row=i_f, column=2).value
            if empty_count_f >= 10:
                break
            if para_n and para_v:
                para_dict_f[para_n] = para_v
            else:
                empty_count_f += 1
        return para_dict_f
    else:
        wb = openpyxl.load_workbook(excel_path_f, data_only=True)
        ws = wb.worksheets[0]
        para_dict_old = {}
        empty_count_f = 0
        for i_f in range(1, 10000):
            para_n = ws.cell(row=i_f, column=1).value
            para_v = ws.cell(row=i_f, column=2).value
            if empty_count_f >= 10:
                break
            if para_n and para_v:
                para_dict_old[para_n] = i_f
            else:
                empty_count_f += 1
        for item_f, value_f in data_f.items():
            if item_f in para_dict_old.keys():
                ws.cell(row=para_dict_old[item_f], column=2, value=value_f)
        wb.save(excel_path_f)
# Conduct EXE file or Open doc in Explorer
def command_conduct(prompt_f):
    if is_url(prompt_f):
        chrome_cmd = "start msedge --new-window " + prompt_f
        subprocess.Popen(chrome_cmd, shell=True)
    elif os.path.exists(prompt_f):
        subprocess.Popen(["explorer", prompt_f])
# To remove overlap of multi calsses, remain elements which rank upper
def rm_dp_list_elem(list_f, reverse=True):
    a_no_duplicates = []
    if reverse:
        for i in reversed(list_f):
            if i not in a_no_duplicates:
                a_no_duplicates.insert(0, i)
    else:
        for i in list_f:
            if i not in a_no_duplicates:
                a_no_duplicates.append(i)
    return a_no_duplicates
# Sorted Association Words according to the location of prompt
def dir_name_sort(dir_list_f, prompt_f):
    order_list = []
    for dir_f in dir_list_f:
        try:
            order_list.append([dir_f.index(prompt_f), dir_f])
        except Exception:
            order_list.append([999999, dir_f])
    return [x[1] for x in sorted(order_list, key=lambda x:x[0])]
# resize geometry
def resize(data_f, para_dict_f, type_f):
    data_f = copy.deepcopy(data_f)
    scrx = para_dict_f['screen'][0]
    scry = para_dict_f['screen'][1]
    res_x = para_dict_f['res_x']
    res_y = para_dict_f['res_y']
    gap = para_dict_f['gap']
    het = para_dict_f['het']
    radius = para_dict_f['srh_r']
    win_x = para_dict_f['main_window'][0]
    win_y = para_dict_f['main_window'][1]
    if type_f == 'font':
        for value_if in data_f.values():
            temp_num = value_if['PixelSize']
            num_d = int(temp_num*min(res_x, res_y))
            value_if['PixelSize'] = num_d
            
    elif type_f == 'window':
        for dict_if in data_f.values():
            for key_f, value_f in dict_if.items():
                if not value_f:
                    continue
                else:
                    x, y, w, h = value_f
                    dict_t = dict()
                    exec(f'x_t=int({x})\ny_t=int({y})\nw_t=int({w})\nh_t=int({h})', locals(), dict_t)
                    dict_if[key_f] = [dict_t['x_t'], dict_t['y_t'], dict_t['w_t'], dict_t['h_t']]
    return data_f
# search for target path
def path_trans(path_dict_f, work_dir_f):
    for window_ni, dict_if in path_dict_f.items():
        for key_if, value_if in dict_if.items():
            if not value_if:
                path_t = glob(os.path.join(work_dir_f, window_ni, f"{key_if}.*"))
                if len(path_t) == 1:
                    path_dict_f[window_ni][key_if] = path_t[0]
                else:
                    print(f"{window_ni}-{key_if}: No Target Detected!")
            else:
                path_t = glob(os.path.join(work_dir_f, window_ni, value_if, "*"))
                if path_t:
                    path_dict_f[window_ni][key_if] = path_t
                else:
                    print(f"{window_ni}-{value_if}: No Target Detected!")
    return path_dict_f
# repalce nan with None
def nan_to_sign(list_f):
    return ['~' if pd.isna(i) else str(i) for i in list_f]

# words match method
class Associate:
    def __init__(self, program_info_df, ouput_num_f):
        self.num = ouput_num_f
        self.df = program_info_df
        self.names = nan_to_sign(list(self.df['Name'].values))
        self.ch_names = nan_to_sign(list(self.df['Chinese Name'].values))
        self.description = nan_to_sign(list(self.df['Description'].values))
        self.exe_path = nan_to_sign(list(self.df['EXE Path'].values))
    
    # To Associate subdirectory of certain path
    def path(self, prompt_f):
        if not prompt_f:
            return []
        # prompt is a specific file path
        if os.path.exists(prompt_f) and os.path.isfile(prompt_f):
            return []
        # prompt is a directory
        if os.path.exists(prompt_f):
            return os.listdir(prompt_f)
        # prompt is a unfilled diretory
        sup_path = os.path.dirname(prompt_f)
        dir_name_n = prompt_f.split('\\')[-1]
        if not os.path.exists(sup_path):
            return []
        else:
            list_1f = sorted([dir_if for dir_if in os.listdir(sup_path) if dir_name_n in dir_if], key=lambda x:x.index(dir_name_n))
            list_2f = sorted([dir_if for dir_if in os.listdir(sup_path) if dir_name_n.lower() in dir_if.lower()], key=lambda x:(x.lower()).index(dir_name_n.lower()))
            return rm_dp_list_elem(list_1f+list_2f, reverse=False)
    
    # To associate a programme name of prompt
    def name(self, prompt_f):
        if not prompt_f:
            return []
        try:
            output_name = sorted([name_i for name_i in self.names if prompt_f.lower() in name_i.lower()], key=lambda s: (s.lower().index(prompt_f.lower()))/len(s))
            output_chname = sorted([ch_name_i for ch_name_i in self.ch_names if prompt_f.lower() in ch_name_i.lower()], key=lambda s: (s.lower().index(prompt_f.lower()))/len(s))
            output_des = [self.names[self.description.index(dsp_name_i)] for dsp_name_i in self.description if prompt_f.lower() in dsp_name_i.lower()]
            output_chname_t = [i for i in output_chname if self.names[self.ch_names.index(i)] not in output_name]
            output_des_t = [i for i in output_des if self.ch_names[self.names.index(i)] not in output_chname]
            return rm_dp_list_elem(output_name+output_chname_t+output_des_t, reverse=False)[0:self.num]
        except Exception as e:
            return []
    
    # To associate programmes with multiple prompts
    def multi_pro_name(self, prompt_list_f):
        output_0 = list_overlap([self.name(prompt_if)[0] for prompt_if in prompt_list_f])
        output_1 = [self.names[(self.ch_names).index(ch_name_i)] for ch_name_i in list_overlap([self.name(prompt_if)[1] for prompt_if in prompt_list_f])]
        output_2 = [self.names[(self.description.index(dsp_name_i))] for dsp_name_i in list_overlap([self.name(prompt_if)[2] for prompt_if in prompt_list_f])]
        return (output_0+output_1+output_2)[0:self.num]
    
    # Automatically complete PATH
    def fill_path(self, prompt_f):
        # if prompt is a exsisting path
        if os.path.exists(prompt_f):
            if len(os.listdir(prompt_f))==1:
                return os.path.join(prompt_f, os.listdir(prompt_f)[0])
            else:
                return None
        else:
            sup_path = os.path.dirname(prompt_f)
            if not os.path.exists(sup_path):
                return None
            else:
                name_wait_list = [name_if for name_if in os.listdir(sup_path) if (name_if).lower().startswith((prompt_f.split('\\')[-1]).lower())]
                if len(name_wait_list) == 1:
                    return os.path.join(sup_path, name_wait_list[0])
                else:
                    return None
    
    # Automatically complete Name
    def fill_name(self, prompt_f):
        output_0 = [name_i for name_i in self.names if name_i.lower().startswith(prompt_f.lower())]
        output_1 = [ch_name_i for ch_name_i in self.ch_names if ch_name_i.lower().startswith(prompt_f.lower())]
        if len(output_0) == 1:
            return output_0[0]
        elif len(output_1) == 1:
            return output_1[0]


# Main Window Set
class SuperLauncher(QMainWindow):
    def __init__(self,
                num_set_f, 
                name_set_f,
                func_set_f,
                font_set_f,
                path_set_f,
                workdir_f,
                window_size_set_ori_f,
                size_set_f,
                app_f,
                media_path_set_f,
                media_font_set_f):
        # call __init__ of QMainWinodw
        super(SuperLauncher, self).__init__()
        # load input parameters
        self.num_set = num_set_f
        self.name_set = name_set_f
        self.func_set = func_set_f
        self.font_set = font_set_f
        self.path_set = path_set_f
        self.workdir = workdir_f
        self.window_size_set_ori = window_size_set_ori_f
        self.size_set = size_set_f
        self.shortcut_df = excel_to_df(self.path_set['Shortcut Setting']['shortcut_set'], region='A:C')
        self.app = app_f
        self.media_path_set = media_path_set_f
        self.media_font_set = media_font_set_f
        
        # for mouse click enven judge
        self.drag_position = None
        self.edge_drag = None
        self.edge_threshold = self.size_set['gap']//3.5
        
        # unpack some essential input parameters
        self.associate = self.func_set['associate_list']
        self.associate_num = self.num_set['associate_list']
       
        # define preload data
        self.associa_l = []
        self.window_size_set = resize(self.window_size_set_ori, 
                                      self.size_set, 
                                      'window') # calculate geometry parameter of windows
        self.window_size_set_ori_data = copy.deepcopy(self.window_size_set)
        self.opto_para_dict = opto_excel_process(self.path_set['Opto Setting']['opto_set'])
        self.color_dict = {'common':QColor("#4A4A4A"),
                               'user':QColor('#2F76D5'),
                               'assistant':QColor("#EA124F"),
                               'command':QColor("#EA124F"),
                               'opto':QColor('#F1F1F1')}
        self.enrty_effect_dict= {}
        
        # All Mode Input
        self.launcher_input = []
        self.searcher_input = []
        self.gpt_input = []
        self.opto_input = []
        self.media_input = []
        self.input_record = {'Launcher':self.launcher_input,
                             'Searcher':self.searcher_input,
                             'GPT':self.gpt_input,
                             'Opto':self.opto_input,
                             'Media':self.media_input}
        self.input_record_index = -1    # for up、down keypress 
        
        
        # gpt dialog load
        self.Mistral_dia = []
        self.CodeLlama_dia = []
        self.Llama2_dia = []
        self.dialogs = {'Mistral-7B-Instruct-v0.2':self.Mistral_dia,
                        'CodeLlama-7b-Instruct-hf':self.CodeLlama_dia,
                        'Llama2-2-7b-chat-hf':self.Llama2_dia
                        }
        # gpt box screen print record
        self.Mistral_output = ''
        self.CodeLlama_output = ''
        self.Llama2_output = ''
        self.gpt_output = {'Mistral-7B-Instruct-v0.2':self.Mistral_output,
                        'CodeLlama-7b-Instruct-hf':self.CodeLlama_output,
                        'Llama2-2-7b-chat-hf':self.Llama2_output
                        }
        
        # opto box screen print record
        self.opto_box_record = ''
        self.opto_open_list = []
        
        # state tag data
        self.state = 'Finished'
        self.model = 'Launcher Mode'
        
        
        
        # define layout
        self.layout_up = QVBoxLayout()
        
        # init window defined in other class
        self.srh_icon = ToggleIconButton(self)
        self.init = Initiate(self)
        self.shortcut_button = Shortcuts(self)
        self.hotkey_init()
        
        # window arrange before launch
        # central_widget.setLayout(self.layout)
        self.setLayout(self.layout_up)  
        self.installEventFilter(self)    # Install the event filter to capture Tab key
        self.raise_()   # Display on the Top Level
        self.show()
    # For mouse control
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            if self.isNearEdge(event.pos()):
                self.edge_drag = event.pos()
                self.original_geometry = self.geometry()
                event.accept()
            else:
                x, y, w, h = self.get_geometry(self.switch_button)
                if event.y() >y+h or event.x() > x+w:
                    self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
                    event.accept()
        else:
            super().mousePressEvent(event)
    def mouseMoveEvent(self, event):
        if self.edge_drag:
            self.resizeWindow(event.pos())
        elif self.drag_position and event.buttons() == Qt.LeftButton:
            self.move(event.globalPos() - self.drag_position)
            event.accept()
        else:
            super().mouseMoveEvent(event)
    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drag_position = None
            self.edge_drag = None
            event.accept()
        else:
            super().mouseReleaseEvent(event)
    def isNearEdge(self, pos):
        x, y, w, h = self.get_geometry(self)
        return (
            pos.x() < self.edge_threshold or
            pos.x() > w - self.edge_threshold or
            pos.y() < self.edge_threshold or
            pos.y() > h - self.edge_threshold
        )
    def resizeWindow(self, pos):
        diff = pos - self.edge_drag
        if abs(diff.x()) > abs(diff.y()):
            new_width = max(self.original_geometry.width() + diff.x(), self.minimumWidth())
            self.setGeometry(self.original_geometry.x(), self.original_geometry.y(), new_width, self.original_geometry.height())
        else:
            new_height = max(self.original_geometry.height() + diff.y(), self.minimumHeight())
            self.setGeometry(self.original_geometry.x(), self.original_geometry.y(), self.original_geometry.width(), new_height)
    # paint background
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # 绘制带有圆角和渐变的窗口背景
        gradient = QLinearGradient(0, 0, 0, self.height())
        ## for unlinear change of background color
        # gradient.setColorAt(0, QColor(23, 139, 122))
        # gradient.setColorAt(0.4, QColor(133, 231, 214))
        # gradient.setColorAt(0.6, QColor(188, 245, 231))
        # gradient.setColorAt(1, QColor(243, 255, 252))
        gradient.setColorAt(0, QColor(113, 148, 139))
        gradient.setColorAt(1, QColor(162, 245, 224))
        painter.setBrush(gradient)
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(self.rect(), 20, 20)
    # To Inhibit Focus Switch when pressing TAB
    def focusNextPrevChild(self, next):
        # Override to prevent normal focus change on Tab key press
        return False
    # supervise keyboard input
    def eventFilter(self, obj, event):
        mode_cat_f = self.clarify_mode(self.get_mode())
        text_n = self.get_input()
        list_temp = self.input_record[mode_cat_f] + [text_n]
        modifiers = QApplication.keyboardModifiers()
        # switch model event filter

        if event.type() == QEvent.KeyPress:
            if modifiers == Qt.ControlModifier:
                return super().eventFilter(obj, event)
            #     # if event.key() == Qt.Key_Tab:
            #     #     # Handle Tab key press with Ctrl
            #     #     self.model_switch((self.switch_button.currentIndex()+1)%len(self.name_set['mode_list']))
            #     # elif event.key() == Qt.Key_QuoteLeft:
            #     #     index_t = self.switch_button.currentIndex()-1
            #     #     index_t = index_t if index_t > -1 else index_t + len(self.name_set['mode_list'])
            #     #     self.model_switch(index_t)
            #     if event.key() >= Qt.Key_0 and event.key() <= Qt.Key_9:
            #         number = event.key() - Qt.Key_0 -1
            #         try:
            #             tar_mode = self.name_set['mode_list'][number]
            #         except Exception:
            #             tar_mode = None
            #         if tar_mode:
            #             self.model_switch(number)
            
            # else:
            if event.key() == Qt.Key_Tab:
                # Handle Tab key press
                self.tab_effect()
            elif event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
                self.confirm_action()
            elif event.key() == Qt.Key_Up:
                if self.input_record_index-1 < -len(list_temp):
                    pass
                else:
                    self.input_record_index -= 1
                    self.fill_input_box(list_temp[self.input_record_index])
            elif event.key() == Qt.Key_Down:
                if self.input_record_index+1 > -1:
                    pass
                else:
                    self.input_record_index += 1
                    self.fill_input_box(list_temp[self.input_record_index])

        return super().eventFilter(obj, event)
    ## programm exit
    @staticmethod
    def programm_exit():
        QApplication.instance().quit()
    # programm restart
    def programm_restart(self):
        if hasattr(self, 'player'):
            self.player.save_data()
        restart_program(fr'''"{self.name_set['script_path']}"''')
    # hotkey set
    def hotkey_init(self):
        # model switch hotkey
        previous_hotkey = QShortcut(QKeySequence("Ctrl+TAB"), self)
        func_back = partial(self.model_tp, prompt_f='Backward')
        previous_hotkey.activated.connect(func_back)
        
        next_hotkey = QShortcut(QKeySequence("Ctrl+Q"), self)
        func_for = partial(self.model_tp, prompt_f='Forward')
        next_hotkey.activated.connect(func_for)
       
        for i in range(0, 10):
            tp_hotkey = QShortcut(QKeySequence(f"Ctrl+{i}"), self)
            func_t = partial(self.model_tp, prompt_f=i-1)
            tp_hotkey.activated.connect(func_t)

        # programm exit hotkey
        exit_hotkey = QShortcut(QKeySequence("Ctrl+D"), self)
        exit_hotkey.activated.connect(self.programm_exit)
        # programm restart hotkey
        restart_hotkey = QShortcut(QKeySequence("Ctrl+R"), self)
        restart_hotkey.activated.connect(self.programm_restart)

        ## music player hotkey
        former_song_hotkey = QShortcut(QKeySequence("Ctrl+W"), self)
        former_song_hotkey.activated.connect(self.player.bottom_last_button_effect)

        next_song_hotkey = QShortcut(QKeySequence("Ctrl+E"), self)
        next_song_hotkey.activated.connect(self.player.bottom_next_button_effect)

        play_pause_hotkey = QShortcut(QKeySequence("Ctrl+Space"), self)
        play_pause_hotkey.activated.connect(self.player.bottom_play_button_effect)



    ## hotkey function
    def model_tp(self, prompt_f):
        if type(prompt_f) == int:
            try:
                tar_mode = self.name_set['mode_list'][prompt_f]
            except Exception as e:
                tar_mode = None
            if tar_mode:
                self.model_switch(tar_mode)
        else:
            if prompt_f == 'Forward':
                self.model_switch(self.name_set['mode_list'][(self.switch_button.currentIndex()+1)%len(self.name_set['mode_list'])])
            elif prompt_f == 'Backward':
                index_t = self.switch_button.currentIndex()-1
                index_t = index_t if index_t > -1 else index_t + len(self.name_set['mode_list'])
                self.model_switch(self.name_set['mode_list'][index_t])

    # read content in input box
    def get_input(self):
        current_text_f = str(self.input_box.text())
        return current_text_f.lstrip()
    # get name of the model
    def get_mode(self):
        s_index_f = self.switch_button.currentIndex()
        s_model_f = self.name_set['mode_list'][s_index_f]
        return s_model_f
    @staticmethod
    def get_geometry(window_f):
        geom_f = window_f.geometry()
        return[geom_f.x(), geom_f.y(), geom_f.width(), geom_f.height()]
    # get mode category of mode
    def clarify_mode(self, mode_name_f):
        for mode_class, value_f in self.name_set['mode_clarify'].items():
            if mode_name_f in value_f:
                return mode_class
    # reset geometry of window
    def reset_geometry_all(self, mode_name_f, mode_cat_f):
        self.size_set['main_window'] = [self.get_geometry(self)[-2], self.get_geometry(self)[-1]]
        self.window_size_set = resize(self.window_size_set_ori, self.size_set, 'window')
        if mode_cat_f != 'GPT':
            try:
                self.progress_bar.hide()
            except Exception:
                pass
        else:
            if self.state == 'Working':
                try:
                    self.progress_bar.show()
                except Exception:
                    pass
        for window_name_if in self.name_set['window_name']:
            if window_name_if == 'switch_button':
                self.switch_button.close()
                self.init.switch_button(mode_name_f, mode_cat_f)
                self.switch_button.show()
            elif window_name_if == 'main_window':
                x, y, w, h = self.get_geometry(self)
                x_1, y_1, w_1, h_1 = self.window_size_set_ori_data[mode_cat_f]['main_window']
                self.setGeometry(x+(w-w_1)//2, y+(h-h_1)//2, w_1, h_1)
            else:
                window_size_f = self.window_size_set[mode_cat_f].get(window_name_if)
                if not window_size_f:
                    exec(f'self.{window_name_if}.hide()', locals())
                    exec(f"self.{window_name_if}.setVisible(False)", locals())
                else:
                    window_size_f = tuple(window_size_f)
                    exec(f"self.{window_name_if}.setGeometry(*window_size_f)", locals())
                    exec(f'self.{window_name_if}.show()', locals())
                    exec(f"self.{window_name_if}.setVisible(True)", locals())
                    if window_name_if in ['opto_icon', 'shortcut_entry']:
                        exec(f'self.{window_name_if}.raise_()', locals())
    # change color
    def set_color(self, window_f, color_f, name_f='common'):
        format_f = QTextCharFormat()
        format_f.setForeground(color_f)
        if name_f=='user' or name_f=='assistant' or name_f == 'command':
            format_f.setFontWeight(80)  # 设置为粗体
            format_f.setFontFamily('Jetbrains Mono')  # 设置字体
        elif name_f == 'input':
            format_f.setFontWeight(70)  # 设置为粗体
        window_f.setCurrentCharFormat(format_f)
    # gpt box print
    def gpt_print(self, text_f, model_name_f):
        gpt_name_s = model_name_f.split("-")[0]
        split_parts = [i_f for i_f in re.split(rf'(User:|Assistant:)', text_f) if i_f]
        
        for text_if in split_parts:
            if 'User:' == text_if:
                cursor = self.gpt_box.textCursor()
                cursor.movePosition(QTextCursor.End)
                text_format = QTextCharFormat()
                text_format.setForeground(self.color_dict['user'])
                cursor.insertText(text_if, text_format)
            elif 'Assistant:' == text_if:
                cursor = self.gpt_box.textCursor()
                cursor.movePosition(QTextCursor.End)
                text_format = QTextCharFormat()
                text_format.setForeground(self.color_dict['assistant'])
                cursor.insertText(text_if, text_format)
            else:
                cursor = self.gpt_box.textCursor()
                cursor.movePosition(QTextCursor.End)
                text_format = QTextCharFormat()
                text_format.setForeground(self.color_dict['common'])
                cursor.insertText(text_if, text_format)
    # opto box print
    def opto_print(self, text_f):
        match_f = re.match(r'(Command@\d+:)\s*(.*)', text_f)
        if match_f:
            self.set_color(self.opto_box, self.color_dict['command'], 'command')
            self.opto_box.insertPlainText(match_f.group(1))
            self.set_color(self.opto_box, self.color_dict['opto'], 'input')
            self.opto_box.insertPlainText(match_f.group(2)+'\n')
        else:
            self.set_color(self.opto_box, self.color_dict['opto'])
            self.opto_box.insertPlainText(text_f+'\n')
    
    # to update associate_words
    def update_associated_words(self):
        current_text = self.get_input()
        if self.replace_path(current_text):
            current_text = self.get_input()
        self.associate_list.clear()
        if is_path(current_text):
            matching_words = self.associate.path(current_text)
        else:
            if ';' not in current_text:
                matching_words = self.associate.name(current_text)
            else:
                matching_words = self.associate.multi_pro_name(current_text.split(';'))

        for word in matching_words:
            item = QListWidgetItem(' '+word)
            self.associate_list.addItem(item)
        try:
            font_metrics = QFontMetrics(self.associate_list.font())
            width_f = max([font_metrics.boundingRect(' '+word).width() + 60 for word in matching_words])
            width_f = min(width_f, (self.window_size_set['Launcher']['input_box'][-2]-self.size_set['gap'])//2)
            self.associate_list.setMinimumWidth(width_f)
            self.associate_list.setMaximumWidth(width_f)
        except Exception:
            pass
    # To respond to TAB KeyPress
    def tab_effect(self):
        mode_name_f = self.get_mode()
        if 'Launcher' not in mode_name_f:
            return
        current_text = self.get_input()
        obj_t = self.associate.fill_path(current_text) if is_path(current_text) else self.associate.fill_name(current_text)
        if obj_t:
            self.fill_input_box(obj_t)
    # The function to write in input box
    def fill_input_box(self, item):
        item_f = item if isinstance(item, str) else str(item.text())
        self.input_box.setText(item_f)
        self.input_box.setFocus()
    # The funtion to update input when clicking on associate list
    def click_update(self, item_f):
        current_text = self.get_input()
        if not is_path(current_text):
            target_f = str(item_f.text()).lstrip()
            self.fill_input_box('')
            if target_f in self.associate.names:
                command_temp = self.associate.exe_path[self.associate.names.index(target_f)]
            elif target_f in self.associate.ch_names:
                command_temp = self.associate.exe_path[self.associate.ch_names.index(target_f)]
            else:
                command_temp = target_f
            worker_process = multiprocessing.Process(target=command_conduct, args=(command_temp,))
            worker_process.start()
            return
        if os.path.exists(current_text):
            self.fill_input_box(os.path.join(current_text, str(item_f.text().lstrip())))
        else:
            self.fill_input_box(os.path.join(os.path.dirname(current_text), str(item_f.text().lstrip())))  
        self.input_box.setFocus()
    # gpt function for threading
    def gpt(self, input_f, model_name_ff, progress_bar_f):
        if not hasattr(self, 'hugface'):
            self.hugface = HuggingFaceGPT()
        try:
            answer_f = self.hugface.chat(input_f)
        except Exception as e:
            answer_f = 'ERROR! ' + str(e)
        self.answer = True
        self.gpt_print(f'Assistant: {answer_f.rstrip()}\n\n', model_name_ff)
        self.gpt_output[model_name_ff] = self.gpt_box.toPlainText()
        time.sleep(0.6)
        self.state = 'Finished'
        progress_bar_f.close()
    # optogenetic sysytem reaction func
    def opto(self, opto_input_f):
        para_dict = copy.deepcopy(self.opto_para_dict)
        output_f = opto_main(para_dict, opto_input_f)
        if isinstance(output_f, str):
            return output_f
        elif isinstance(output_f, list):
            self.opto_open_list = output_f[1]
            return output_f[0]
        else:
            return 'Unkown Output Format!'
    # conduct function
    def confirm_action(self):
        input_con_f = self.get_input()
        if not input_con_f:
            return
        self.input_record_index = -1
        mode_name_f = self.get_mode()
        mode_cat_f = self.clarify_mode(mode_name_f)
        self.input_record[mode_cat_f].append(input_con_f)
        self.fill_input_box('')
        pattern_t = r'^resize\((\d+),\s*(\d+)\)$'
        if input_con_f == 'resize()':
            x_1, y_1, w_1, h_1 = self.window_size_set_ori_data[mode_cat_f]['main_window']
            self.resize(w_1, h_1)
            return
        elif input_con_f == 'restart()':
            self.programm_restart()
        elif re.match(pattern_t, input_con_f):
            match_f = re.match(pattern_t, input_con_f)
            w_t, h_t = int(match_f.group(1)), int(match_f.group(2))
            self.resize(w_t, h_t)
            return
        elif input_con_f == 'exit()':
            self.programm_exit()
        if mode_cat_f == 'Launcher':
            if is_path(input_con_f):
                command_temp = input_con_f.replace('/', '\\')
            elif is_url(input_con_f):
                command_temp = input_con_f
            else:
                if input_con_f in self.associate.names:
                    command_temp = self.associate.exe_path[self.associate.names.index(input_con_f)]
                elif input_con_f in self.associate.ch_names:
                    command_temp = self.associate.exe_path[self.associate.ch_names.index(input_con_f)]
                else:
                    command_temp = ''
            worker_process = multiprocessing.Process(target=command_conduct, args=(command_temp,))
            worker_process.start()
        elif mode_cat_f == 'Searcher':
            browser_name = self.srh_icon_paths[self.current_index].split('\\')[-1]
            if 'bing' in browser_name:
                url_f = fr"https://www.bing.com/search?q={input_con_f}"
            elif 'baidu' in browser_name:
                url_f = fr"baidu.com/s?ie=utf-8&f=8&rsv_bp=1&rsv_idx=1&tn=baidu&wd={input_con_f}&fenlei=256&rsv_pq=0xd73949d80006e0bb&rsv_t=3c845DZ6fB%2BS%2BQRyuXkVEjoJ6p2Svo2fkc3GilNyKwzIj9YJnRpHptz%2BVKfn&rqlang=en&rsv_enter=1&rsv_dl=tb&rsv_sug3=9&rsv_sug1=8&rsv_sug7=101&rsv_sug2=0&rsv_btyp"
            elif 'google' in browser_name:
                url_f = fr"https://www.google.com.hk/search?q={input_con_f}&newwindow=1&sca_esv=600400644&source=hp&ei=hmCuZZuoJejGkPIPmNikuAM&iflsig=ANes7DEAAAAAZa5ulnJYxuARlbG1biZ4CdvLcrbPo0hB&ved=0ahUKEwibg_jCgPGDAxVoI0QIHRgsCTcQ4dUDCA0&uact=5&oq=%E4%BD%A0%E5%A5%BD&gs_lp=Egdnd3Mtd2l6IgbkvaDlpb0yBRAAGIAEMgUQLhiABDIFEC4YgAQyBRAAGIAEMgUQLhiABDILEC4YgAQYxwEY0QMyBRAAGIAEMgUQLhiABDIFEC4YgAQyCBAuGIAEGNQCSM4pUNkVWIEmcAN4AJABAJgBtQKgAfQLqgEDMy01uAEDyAEA-AEBqAIAwgILEC4YgAQYxwEYrwE&sclient=gws-wiz"
            else:
                url_f = ''
            worker_process = multiprocessing.Process(target=command_conduct, args=(url_f,))
            worker_process.start()
        elif mode_cat_f == 'GPT':
            if input_con_f == 'clear()':
                self.gpt_box.clear()
                self.dialogs[mode_name_f].clear()
                self.gpt_output[mode_name_f] = None
                if hasattr(self, 'hugface'):
                    self.hugface.dialog.clear()
                return None
            self.dialogs[mode_name_f].append({'role':'user', 'content':input_con_f})
            self.gpt_print(f'User: {input_con_f}\n', mode_name_f)
            if not hasattr(self, 'answer'):
                self.answer = False
            self.progress_bar = Progress_bar(self, 8)
            gpt_thread = threading.Thread(target=self.gpt, args=(input_con_f, mode_name_f, self.progress_bar))
            gpt_thread.start()
            self.progress_bar.setGeometry(*tuple(self.window_size_set[mode_cat_f]['input_box']))
            self.state = 'Working'
            self.progress_bar.show()
        elif mode_cat_f == 'Opto':
            if input_con_f == 'clear()':
                self.opto_box.clear()
                self.opto_box_record.clear()
                return
            self.opto_print(f'Command@{len(self.input_record["Opto"])}: {input_con_f}\n')
            match_f = re.match(r'([Oo][Pp])\s+(\d+)', input_con_f)
            if match_f:
                index_f = match_f.group(2)
                if index_f < len(self.opto_open_list):
                    path_if = self.opto_open_list[index_f]
                    subprocess.Popen(['explorer', path_format(path_if)])
                    self.opto_print(f'{path_if} IS Opened!')
                else:
                    self.opto_print('Index Out Of Range!')
                return
            self.opto_print(self.opto(input_con_f))
            return
            
    # add input load
    def add_input(self, prompt_f):
        mode_nf = self.get_mode()
        mode_cat_f = self.clarify_mode(mode_nf)
        if mode_cat_f == 'Launcher':
            self.launcher_mode_input.append(prompt_f)
        elif mode_cat_f == 'Searcher':
            self.searcher_mode_input.append(prompt_f)
        elif mode_cat_f == 'GPT':
            self.gpt_mode_input.append(prompt_f)
    
    # refresh associate list and fill in object once typing in 
    def associate_refresh(self):
        init_input = ''
        while True:
            model_name = self.get_mode()
            if  'Launcher' not in model_name:
                time.sleep(0.015)
                continue
            input_now = self.get_input()
            if input_now == init_input:
                time.sleep(0.015)
                continue
            self.update_associated_words()
            init_input = input_now
    # Model Switch Function
    def model_switch(self, index):
        model_tar = self.switch_button.itemText(index) if isinstance(index, int) else index
        mode_cat_f = self.clarify_mode(model_tar)
        if model_tar == self.model:
            return
        self.input_record_index = -1
        self.model = model_tar
        # refresh button content and size
        if mode_cat_f == 'GPT':
            self.gpt_box.clear()
            if self.gpt_output[model_tar]:
                self.gpt_print(self.gpt_output[model_tar], str(model_tar))
        self.reset_geometry_all(model_tar, mode_cat_f)
    # resize gpu_box
    def resizeEvent(self, event):
        # 当窗口大小改变时，调整按钮的大小
        new_size_f = event.size()
        gap, het = self.size_set['gap'], self.size_set['het']
        win_xf, win_yf = new_size_f.width(), new_size_f.height()
        self.size_set['main_window'] = [win_xf, win_yf]
        self.window_size_set = resize(self.window_size_set_ori, self.size_set, 'window')
        self.input_box.resize(win_xf-2*gap, het)
        # win_y-2.8*gap-2*het
        
        self.gpt_box.resize(win_xf-2*gap, int(win_yf-2.8*gap-2*het))
        self.opto_box.resize(win_xf-2*gap, int(win_yf-2.8*gap-2*het))
        self.associate_list.resize(win_xf-2*gap, int(win_yf-2.8*gap-2*het))
        
        x, y, w, h = self.get_geometry(self.srh_icon)
        self.srh_icon.setGeometry(win_xf//2-w//2, y, w, h)
        
        x_1, y_1, w_1, h_1 = self.get_geometry(self.opto_icon)
        self.opto_icon.setGeometry(win_xf-gap-het, y_1, w_1, h_1)
        
        x_1, y_1, w_1, h_1 = self.get_geometry(self.shortcut_entry)
        self.shortcut_entry.setGeometry(win_xf-gap-het, y_1, w_1, h_1)
        
        x_1, y_1, w_1, h_1 = self.get_geometry(self.top_widget)
        self.top_widget.setGeometry(win_xf-gap//2-4*het, y_1, w_1, h_1)
        
        self.shortcut_button.close()
        self.shortcut_button = Shortcuts(self)
        if self.get_mode() != 'Launcher Mode':
            self.shortcut_button.hide()
        
        try:
            self.shortcut_setting.raise_()
        except Exception:
            pass
    # replace e\ with e:\
    def replace_path(self, text_f):
        if self.model != 'Launcher Mode':
            return False
        pattern_f = r"([a-zA-Z])\\"
        match_f = re.match(pattern_f, rf'{text_f}')
        if not match_f:
            return False
        letter_f = match_f.group(1)
        self.fill_input_box(f'{letter_f}:\\')
        return True
    # Opto Settings Window
    def opto_setting_entry(self):
        self.opto_setting = Opto_settings(self)
        self.opto_setting.show()
    # open shortcut setting windw
    def shortcut_setting_launch(self):
        self.shortcut_setting = Shortcuts_setting(self)
        self.shortcut_setting.show()
    # animation define function
    def animations(self, button_f, size_ori, size_dst, func_f):
        self.animation1 = QPropertyAnimation(self, targetObject=button_f, propertyName=b'iconSize')
        self.animation1.setDuration(80)
        self.animation1.setStartValue(QSize(size_ori, size_ori))
        self.animation1.setEndValue(QSize(size_dst, size_dst))
        self.animation1.finished.connect(func_f)
        self.animation1.start()
    # button click total func
    def button_flash(self):
        button_f = self.sender()
        func_f = self.enrty_effect_dict[button_f.objectName()]
        size_ori = button_f.iconSize().width()
        size_dst = int(button_f.iconSize().width()*0.4)
        self.animations(button_f, size_ori, size_dst, lambda:self.animations(button_f, size_dst, size_ori, func_f))

    
    # for top button control
    def maximum(self):
        x, y, w, h = self.get_geometry(self)
        x_ori, y_ori, w_ori, h_ori = self.window_size_set_ori_data[self.clarify_mode(self.get_mode())]['main_window']
        scrx, scry = self.size_set['screen']
        if w==scrx and h==scry:
            self.max_button.setIcon(QIcon(self.path_set['main_window']['maximum']))
            self.setGeometry(x_ori, y_ori, w_ori, h_ori)
        else:
            self.max_button.setIcon(QIcon(self.path_set['main_window']['middle']))
            self.setGeometry(0, 0, scrx, scry)
            

# inhibit focus move when pressing TAB
class CustomListWidget(QListWidget):
    def __init__(self, parent=None):  
        super(CustomListWidget, self).__init__(parent)

    def focusNextPrevChild(self, next):
        # Override to prevent normal focus change
        return True
# window part initiate            
class Initiate:
    def __init__(self, upper_class):
        self.up = upper_class

        self.up.top_layout = QHBoxLayout()
        self.main_window()
        self.associate_list()
        self.gpt_box()

        self.up.gpt_box.hide()
        self.opto_box()
        self.up.opto_box.hide()
        self.input_box()
        self.switch_button()

        self.opto_icon()
        self.up.opto_icon.hide()
        self.shortcut_entry()
        self.top_buttons()
        self.up.player = MusicPlayer(self.up)
        atexit.register(self.up.player.save_data)
        self.up.player.hide()
        self.up.top_widget = QWidget()
        self.up.top_widget.setLayout(self.up.top_layout)
        self.up.top_widget.setGeometry(self.up.size_set['main_window'][0]-self.up.size_set['gap']//2-self.up.size_set['het']*4, 
                                       self.up.size_set['gap']//2, 
                                       self.up.size_set['het']*4, 
                                       int(1.4*self.up.size_set['het']))
        self.up.layout().addWidget(self.up.top_widget)
        self.up.top_widget.show()
        self.up.top_widget.raise_()
        
    # define main_window
    def main_window(self):
        # self.setWindowFlags(Qt.FramelessWindowHint)     # hide top lane
        self.up.setWindowFlags(self.up.windowFlags() | Qt.FramelessWindowHint)
        self.up.setAttribute(Qt.WA_TranslucentBackground)
        # set background
        # background_image_path = self.up.path_set['background_path']
        # background_image = QPixmap(background_image_path)
        # palette = self.up.palette()
        # palette.setBrush(QPalette.Window, background_image)
        # self.up.setPalette(palette)
        
        # initialize the main window
        self.up.setWindowIcon(QIcon(self.up.path_set['main_window']['taskbar_icon']))
        self.up.setWindowTitle(self.up.name_set['main_window'])
        main_window_set = tuple(self.up.window_size_set['Launcher']['main_window'])
        self.up.setGeometry(*main_window_set)    # [x, y, w, h]
        # self.up.setFixedSize(main_window_set[-2], main_window_set[-1])
    # associate list define
    def associate_list(self):
        self.up.associate_list = CustomListWidget(self.up)
        self.up.associate_list.itemClicked.connect(self.up.click_update)
        
        self.up.associate_list.setFont(font_get(self.up.font_set['associate_list']))
        associate_list_set = tuple(self.up.window_size_set['Launcher']['associate_list'])
        self.up.associate_list.setGeometry(*associate_list_set)
        self.up.associate_list.setStyleSheet("""
        QListWidget {
                border: 3px solid black; /* 设置边框 */
            }
            
        QListView {
            background: transparent;  /* 设置背景颜色 */
            border: 10px solid #CCCCCC;  /* 设置边框 */
            border-radius: 20px;  /* 设置边角弧度 */
        }

        QListView::item {
            padding:1px;  /* 设置每个项的内边距 */
            background-color: #FFFFFF;  /* 设置背景颜色 */
            border-radius: 10px;  /* 设置边角弧度 */
        }
        
        QListView::item:hover {
            background-color: #FFA98F;  /* 设置悬停时的背景颜色 */
        }
        
        QListView::item:selected {
            background-color: #14B699;  /* 设置选中项的背景颜色 */
            color: white;  /* 设置选中项的文字颜色 */
        }
        """)
        self.up.associate_list.verticalScrollBar().setStyleSheet("""
        QScrollBar:vertical {
            background: transparent;
            width: 25px;
            margin: 5px;
        }

        QScrollBar::handle:vertical {
            background: #32CC99;
            min-height: 80px;
            border-radius: 7px;
        }

        QScrollBar::sub-line:vertical,
        QScrollBar::add-line:vertical,
        QScrollBar::sub-page:vertical,
        QScrollBar::add-page:vertical {
            background: none;
        }
        """)
        self.up.associate_list.horizontalScrollBar().setStyleSheet("""
        QScrollBar:horizontal {
            background: transparent;
            height: 30px;
            margin: 3px;
        }

        QScrollBar::handle:horizontal {
            background: #32CC99;
            min-width: 80px;
            border-radius: 10px;
        }

        QScrollBar::sub-line:horizontal,
        QScrollBar::add-line:horizontal,
        QScrollBar::sub-page:horizontal,
        QScrollBar::add-page:horizontal,
        QScrollBar::sub-line:vertical,
        QScrollBar::add-line:vertical,
        QScrollBar::sub-page:vertical,
        QScrollBar::add-page:vertical,
        QScrollBar::sub-line:corner,
        QScrollBar::add-line:corner {
        background: transparent;
        }
        """)
        self.up.associate_list.setSpacing(1)  # 设置 item 之间的间距为 10 像素
        self.up.associate_list.setAttribute(Qt.WA_TranslucentBackground)
    # GPT Dialog Box
    def gpt_box(self):
        # 终端
        self.up.gpt_box = QTextEdit(self.up)
        self.up.gpt_box.setReadOnly(True)
        # gpt_box_set = tuple(self.up.window_size_set['Launcher']['gpt_box'])
        self.up.gpt_box.setGeometry(0,0,1,1)
        self.up.gpt_box.setFont(font_get(self.up.font_set['gpt_box']))
        self.up.gpt_box.setStyleSheet("""
            QTextEdit {
                color: #C2BEB9;
                background-color: rgba(255, 255, 255, 50);
                border: 0px solid #1E1F22;
                border-radius: 20px}""")
        self.up.gpt_box.verticalScrollBar().setStyleSheet("""
        QScrollBar:vertical {
            background: transparent;
            width: 10px;
            margin: 0px;
        }

        QScrollBar::handle:vertical {
            background: #ABABA8 ;
            min-height: 20px;
            border-radius: 5px;
        }

        QScrollBar::sub-line:vertical,
        QScrollBar::add-line:vertical,
        QScrollBar::sub-page:vertical,
        QScrollBar::add-page:vertical {
            background: none;
            }
        """)
    # Optoenetic System Dialog Box
    def opto_box(self):
        # 终端
        self.up.opto_box = QTextEdit(self.up)
        self.up.opto_box.setReadOnly(True)
        # opto_box_set = tuple(self.up.window_size_set['Launcher']['opto_box'])
        self.up.opto_box.setGeometry(0,0,1,1)
        self.up.opto_box.setFont(font_get(self.up.font_set['opto_box']))
        self.up.opto_box.setStyleSheet("""
            QTextEdit {
                color: #C2BEB9;
                background-color: #1E1F22;
                border: 1px solid #1E1F22;
                border-radius: 20px}""")
        self.up.opto_box.verticalScrollBar().setStyleSheet("""
        QScrollBar:vertical {
            background: transparent;
            width: 10px;
            margin: 0px;
        }

        QScrollBar::handle:vertical {
            background: #ABABA8 ;
            min-height: 80px;
            border-radius: 5px;
        }

        QScrollBar::sub-line:vertical,
        QScrollBar::add-line:vertical,
        QScrollBar::sub-page:vertical,
        QScrollBar::add-page:vertical {
            background: none;
            }
        """)
        # self.up.opto_box.setStyleSheet("""
        #     QTextEdit {
        #         color: #1E1F22;
        #         background-color: #FFFFFF;
        #         border: 1px solid #FFFFFF;
        #         border-radius: 20px}""")
    # input box define
    def input_box(self, geometry_f=None):
        # define input box
        self.up.input_box = QLineEdit(self.up)
        self.up.input_box.setStyleSheet("border-radius: 20px; padding-left: 20px;")  # smooth four angle
        self.up.input_box.textChanged.connect(self.up.update_associated_words)
        self.up.input_box.returnPressed.connect(self.up.confirm_action)
        self.up.input_box.setFont(font_get(self.up.font_set['input_box']))
        if not geometry_f:
            input_box_set = tuple(self.up.window_size_set['Launcher']['input_box'])
            # Scroll the selection into view
            self.up.input_box.setGeometry(*input_box_set)
        else:
            self.up.input_box.setGeometry(*tuple(geometry_f))

    # switch button define
    def switch_button(self, mode_nam_f='Launcher Mode', mode_cat_f='Launcher'):
        # 创建模式选择的组合框
        self.up.switch_button = QComboBox(self.up)
        for model_if in self.up.name_set['mode_list']:
            self.up.switch_button.addItem(model_if)
        self.up.switch_button.setCurrentIndex(self.up.name_set['mode_list'].index(mode_nam_f))
        self.up.switch_button.currentIndexChanged.connect(self.up.model_switch)  # 连接信号槽
        font_t = font_get(self.up.font_set['switch_button'])
        self.up.switch_button.setFont(font_get(self.up.font_set['switch_button']))
        
        # self.up.blur_effect = QGraphicsBlurEffect()
        # self.up.switch_button.setBlurRadius(10)  # 设置模糊半径
        # self.up.switch_button.setGraphicsEffect(self.blur_effect)
        # 使用 QFontMetrics 计算文本宽度
        font_metrics = QFontMetrics(self.up.switch_button.font())
        w_f = font_metrics.boundingRect(mode_nam_f).width() + 30  # 加上一些额外空间，你可以根据需要修改
        x, y, w, h = self.up.window_size_set[mode_cat_f]['switch_button']
        self.up.switch_button.setGeometry(x, y, w_f, h)
        button_style = """
        QComboBox {
            background-color: rgba(255, 255, 255, 50);
            border: 0px solid #2980b9;
            border-radius: 10px;
            padding: 5px;
            color: white;
        }

        QComboBox::drop-down {
            border: none;  /* 移除下拉箭头的边框 */
        }
        
        QComboBox QAbstractItemView {
        border-radius: 0px;
        background-color:  qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #73B1A7, stop:1 #EDFFFF);
        border: 0px solid #CCCCCC; /* 下拉列表的边框 */
        color: black; /* 下拉列表的文本颜色 */
        selection-background-color: #4CC1B5; /* 选中项的背景颜色 */
        selection-color: black; /* 选中项的文本颜色 */
        }
        
        
        """

        self.up.switch_button.setStyleSheet(button_style)
    # Button to open Opto Para Set
    def opto_icon(self):
        self.up.opto_icon = QToolButton(self.up)
        icon_path = self.up.path_set['Opto Setting']['entry_button']
        self.up.opto_icon.setIcon(QIcon(icon_path))  # 替换成你的图标路径
        x, y, w ,h = self.up.window_size_set['Opto']['opto_icon']
        self.up.opto_icon.setGeometry(*tuple(self.up.window_size_set['Opto']['opto_icon']))
        self.up.opto_icon.setIconSize(QSize(w, h))  # 设置图标大小
        self.up.opto_icon.setFixedSize(w, h)  # 设置按钮大小
        button_name_f = 'opto_entry'
        self.up.opto_icon.setObjectName(button_name_f)
        self.up.enrty_effect_dict[button_name_f] = self.up.opto_setting_entry
        self.up.opto_icon.clicked.connect(self.up.button_flash)
        self.up.opto_icon.setStyleSheet("QToolButton { border: none; background: none; }")  # 设置无边框无背景样式
        self.up.opto_icon.raise_()
    # Button to shortcut setting window
    def shortcut_entry(self):
        self.up.shortcut_entry = QToolButton(self.up)
        icon_path = self.up.path_set['Shortcut Setting']['entry_button']
        self.up.shortcut_entry.setIcon(QIcon(icon_path))  # 替换成你的图标路径
        x, y, w ,h = self.up.window_size_set['Launcher']['shortcut_entry']
        self.up.shortcut_entry.setIconSize(QSize(w, h))  # 设置图标大小
        self.up.shortcut_entry.setFixedSize(w, h)  # 设置按钮大小
        button_name_f = 'shortcut_entry'
        self.up.shortcut_entry.setObjectName(button_name_f)
        self.up.enrty_effect_dict[button_name_f] = self.up.shortcut_setting_launch
        self.up.shortcut_entry.clicked.connect(self.up.button_flash)
        self.up.shortcut_entry.setStyleSheet("QToolButton { border: none; background: none; }")  # 设置无边框无背景样式
        self.up.shortcut_entry.setGeometry(*tuple(self.up.window_size_set['Launcher']['shortcut_entry']))
    
    # top buttons define
    def top_buttons(self):
        het = self.up.size_set['het']
        self.up.max_button = QPushButton()
        self.up.max_button.setIcon(QIcon(self.up.path_set['main_window']['maximum']))
        self.up.max_button.setIconSize(QSize(int(0.6*het), int(0.6*het)))  # 设置图标大小
        self.up.max_button.setFixedSize(het, het)
        self.up.max_button.setStyleSheet('''
            QPushButton { 
                border: none; 
                background: transparent; 
                border-radius: 20px
                }
            
            QPushButton:hover { 
                        background-color: #619F99; 
                }
        ''')  # 设置无边框无背景样式
        self.up.max_button.clicked.connect(self.up.maximum)
        
        self.up.mini_button = QPushButton()
        self.up.mini_button.setIcon(QIcon(self.up.path_set['main_window']['minimum']))
        self.up.mini_button.setIconSize(QSize(int(0.8*het), int(0.8*het)))  # 设置图标大小
        self.up.mini_button.setFixedSize(het, het)
        self.up.mini_button.setStyleSheet('''
            QPushButton { 
                border: none; 
                background: transparent; 
                border-radius: 20px
                }
            
            QPushButton:hover { 
                        background-color: #619F99; 
                }
        ''')  # 设置无边框无背景样式
        self.up.mini_button.clicked.connect(self.up.showMinimized)

        self.up.close_button = QPushButton()
        self.up.close_button.setIcon(QIcon(self.up.path_set['main_window']['close']))
        self.up.close_button.setIconSize(QSize(int(0.6*het), int(0.6*het)))  # 设置图标大小
        self.up.close_button.setFixedSize(het, het)
        self.up.close_button.setStyleSheet('''
            QPushButton { 
                border: none; 
                background: transparent; 
                border-radius: 20px
                }
            
            QPushButton:hover { 
                        background-color: #619F99; 
                }
        ''')  # 设置无边框无背景样式
        self.up.close_button.clicked.connect(self.up.close)
        
        self.up.top_layout.addWidget(self.up.mini_button)
        self.up.top_layout.addWidget(self.up.max_button)
        self.up.top_layout.addWidget(self.up.close_button)
    
# produce progress bar
class Progress_bar(QProgressBar):
    def __init__(self, parent, expected_time):
        self.exp_time = expected_time
        self.up = parent
        super(Progress_bar, self).__init__(parent)
        self.setMaximum(100)
        self.value = 0
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_progress)
        self.timer.start(20)
        

    def update_progress(self):
        if not self.up.answer:
            if self.value <= 95:
                refresh_times = self.exp_time*1000/20
                avg_step = self.maximum()/refresh_times
                amply_size = random.uniform(0.5, 1.5)
                self.value = min(self.value + avg_step*amply_size, 95.5)
                self.setValue(self.value)
            else:
                pass
        else:
            self.up.answer = None
            self.value = 99.9
            self.setValue(self.value)

            
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # 外边框
        border_rect = self.rect().adjusted(0, -1, 0, 0)     # left, up, right, bottom
        # painter.setPen(QPen(QColor("#E74C3C"), 0))  # 设置边框颜色和宽度
        pen = QPen(QColor(0, 0, 0, 0))  # 透明色的 QPen
        pen.setStyle(Qt.NoPen)  # 设置边框样式为无边框
        painter.setPen(pen)
        painter.drawRoundedRect(border_rect, 20, 20)

        # 填充块
        progress_rect = border_rect.adjusted(0, 0, +2, 0)
        progress_width = progress_rect.width() * self.value / self.maximum()
        progress_rect.setWidth(progress_width)
        brush = QBrush(QColor("#10B5A9"))
        painter.setBrush(brush)
        painter.drawRoundedRect(progress_rect, 20, 20)
# seracher engine icon
class ToggleIconButton(QToolButton):
    def __init__(self, parent):
        super(ToggleIconButton, self).__init__(parent)
        self.up = parent
        self.up.srh_icon_paths = self.up.path_set['Searcher Mode']['srh_icon_set']
        self.up.srh_icon_size = (self.up.window_size_set['Searcher']['srh_icon'][-2], 
                                 self.up.window_size_set['Searcher']['srh_icon'][-1])
        self.up.current_index = 0
        self.setIcon(QIcon(self.up.srh_icon_paths[self.up.current_index]))
        self.setIconSize(QSize(*self.up.srh_icon_size))
        self.setCheckable(True)
        self.setStyleSheet("""
        QToolButton {
            background: none;
            border: none;
            padding: 0px;
        }

        QToolButton:checked {
            background: none;
            border: none;
            padding: 0px;
        }
        """)
        self.clicked.connect(self.button_flash)
        self.setGeometry(*tuple(self.up.window_size_set['Searcher']['srh_icon']))
        parent.layout_up.addWidget(self)
        self.setVisible(False)
    
    # animation define function
    def animations(self, button_f, size_ori, size_dst, func_f):
        self.animation1 = QPropertyAnimation(self, targetObject=button_f, propertyName=b'iconSize')
        self.animation1.setDuration(60)
        # self.animation1.setTargetObject(button_f)
        self.animation1.setStartValue(QSize(size_ori, size_ori))
        self.animation1.setEndValue(QSize(size_dst, size_dst))
        self.animation1.finished.connect(func_f)
        self.animation1.start()
    
    def toggle_icon(self):
        self.up.current_index = (self.up.current_index + 1) % len(self.up.srh_icon_paths)
        self.setIcon(QIcon(self.up.srh_icon_paths[self.up.current_index]))
    
    # button click total func
    def button_flash(self):
        button_f = self.sender()
        func_f = self.toggle_icon
        size_ori = button_f.iconSize().width()
        size_dst = int(button_f.iconSize().width()*0.5)
        self.animations(button_f, size_ori, size_dst, lambda:self.animations(button_f, size_dst, size_ori, func_f))

# Optogenetic Parameter Set Window
class Opto_settings(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.up = parent
        self.setWindowFlags(self.windowFlags() | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAutoFillBackground(True)
        self.setWindowIcon(QIcon(self.up.path_set['Opto Setting']['entry_button']))
        self.setWindowTitle('Opto Setting')
        # 参数列表
        self.origin_parameter = copy.deepcopy(self.up.opto_para_dict)
        self.parameters_f = self.up.opto_para_dict
        self.line_edits = {}
        font_metrics = QFontMetrics(font_get(self.up.font_set['opto_setting_button']))
        label_length = max([font_metrics.boundingRect(word).width()+20 
                            for word in self.parameters_f.keys()])
        # 创建布局
        layout = QVBoxLayout()
        title_label = QLabel("Opto Path Settings")
        title_label.setFont(font_get(self.up.font_set['opto_setting_title']))
        title_label.setFixedHeight(int(1.5*self.up.size_set['het']))  # 指定标题高度
        title_label.setStyleSheet("background-color: transparent;")  # 设置标题透明背景
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        # 创建每个参数的输入框、标签和按钮
        for key_f, value_f in self.parameters_f.items():
            param_layout = QHBoxLayout()

            # 标签
            label = QLabel(f"{key_f}")
            label.setFont(font_get(self.up.font_set['opto_setting_button']))
            label.setFixedWidth(label_length)
            label.setAlignment(Qt.AlignCenter)
            param_layout.addWidget(label)

            # 输入框
            line_edit = QLineEdit()
            line_edit.setPlaceholderText(value_f)
            line_edit.setFont(font_get(self.up.font_set['opto_setting_lineedit']))
            self.line_edits[key_f] = line_edit
            line_edit.textChanged.connect(lambda text, key=key_f, le=line_edit: self.check_path_validity(text, key, le))
            param_layout.addWidget(line_edit)

            # 浏览按钮
            browse_button = QPushButton("Browse")
            browse_button.setFont(font_get(self.up.font_set['opto_setting_button']))
            browse_button.clicked.connect(lambda p=key_f, le=line_edit: self.browse_path(p, le))
            param_layout.addWidget(browse_button)

            layout.addLayout(param_layout)
        
        # 创建确认和取消按钮
        confirm_button = QPushButton("Confirm")
        confirm_button.setFont(font_get(self.up.font_set['opto_setting_button']))
        confirm_button.clicked.connect(self.confirm_settings)
        confirm_button.setStyleSheet('''QPushButton {
        background-color: #286FB1;}

        QPushButton:hover {
            background-color: #2CC1E1; /* 设置悬停时的颜色 */
        }''')
        
        
        cancel_button = QPushButton("Cancel")
        cancel_button.setFont(font_get(self.up.font_set['opto_setting_button']))
        cancel_button.clicked.connect(self.cancel_settings)
        cancel_button.setStyleSheet('''QPushButton {
        background-color: #C60618;}

        QPushButton:hover {
            background-color: #FF4A6C; /* 设置悬停时的颜色 */
        }''')
        
        excel_open_button = QPushButton("Config")
        excel_open_button.setFont(font_get(self.up.font_set['opto_setting_button']))
        excel_open_button.clicked.connect(self.open_config_excel)
        
        
        self.setStyleSheet('''
            QLabel {
                color: #333;
                background-color: #44BBB1; /* 背景颜色 */
                padding: 5px; /* 内边距 */
                border-radius: 10px;
            }

            QLineEdit {
                padding: 5px;
                border: 0px solid #999;
                border-radius: 10px;
                alignment: left;
            }

            QPushButton {
                padding: 10px;
                background-color: #44BBB1;
                color: #fff;
                border: none;
                border-radius: 10px;
            }     

            QPushButton:hover {
                background-color: #2CE1B3;
            }

            QPushButton:pressed {
                background-color: #349179;
            }
        ''')
        button_layout = QHBoxLayout()
        button_layout.addWidget(excel_open_button)
        button_layout.addWidget(confirm_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        x, y, w, h = self.up.get_geometry(self.up)
        x_1, y_1, w_1, h_1 = self.up.window_size_set['Opto']['opto_settings']
        self.setGeometry(x+w//2-w_1//2, y+h//2-h_1//2, w_1, h_1)
        self.setLayout(layout)
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # 绘制带有圆角和渐变的窗口背景
        gradient = QLinearGradient(0, 0, 0, self.height())
        gradient.setColorAt(0, QColor(113, 148, 139))
        gradient.setColorAt(1, QColor(162, 245, 224))
        painter.setBrush(gradient)
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(self.rect(), 20, 20)
    # For mouse control
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton:
            self.move(event.globalPos() - self.drag_position)
            event.accept()
    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            event.accept()
    def browse_path(self, param, line_edit):
        # 打开文件对话框，获取文件路径
        file_dialog = QFileDialog()
        if param == 'Excel Path':
            file_path, _ = file_dialog.getOpenFileName(self, "Select File", "", "All Files (*)")
        else:
            file_path = QFileDialog.getExistingDirectory(None, "Select Folder", "", QFileDialog.ShowDirsOnly)
        # 更新输入框的内容
        if file_path:
            line_edit.setText(file_path.replace('/', '\\'))
            # self.parameters_f[param] = file_path.replace('/', '\\')
            

    def confirm_settings(self):
        for key, value in self.parameters_f.items():
            line_edit = self.line_edits[key].text()
            if os.path.isabs(line_edit):
                self.parameters_f[key] = line_edit
        self.up.opto_para_dict = self.parameters_f     
        opto_excel_process(self.up.path_set['Opto Setting']['opto_set'], self.parameters_f)
        self.close()

    def cancel_settings(self):
        self.close()
    
    def open_config_excel(self):
        subprocess.Popen(['explorer', path_format(self.up.path_set['Opto Setting']['opto_set'])])
    def check_path_validity(self, text, key, line_edit):
        text_f = line_edit.text()
        if os.path.exists(text_f) or (not text_f):
            line_edit.setStyleSheet("border: 0px solid #FFFFFF; background-color: #FFFFFF;")  # 路径不存在，变红
        elif os.path.isabs(text_f):
            line_edit.setStyleSheet("border: 0px solid #EAF708; background-color: #FFFFCC;")  # 路径不存在，变红
        else:
            line_edit.setStyleSheet("border: 0px solid #EF3A13; background-color: #FFCCCC;")  # 路径存在，变黄
# Shortcut Button Set
class Shortcuts(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.setParent(parent)
        self.up = parent
        self.df = self.up.shortcut_df
        self.button_list = {}
        self.click_effect = {}
        self.total_layout = QVBoxLayout()
        self.size_f = self.up.window_size_set['Launcher']['shortcut_button']
        self.buttons()
        self.setLayout(self.total_layout)
        self.setGeometry(*tuple(self.size_f))
        self.show()
    def buttons(self):
        obj_num_f = len(self.df)
        hor_num = self.up.num_set['horizontal_button_num']
        ver_num = math.ceil(obj_num_f/hor_num)
        avg_width = self.size_f[-2]/hor_num
        avg_height = self.size_f[-1]/ver_num
        icon_size = int(0.5*min(avg_width, avg_height))
        button_size = int(0.9*min(avg_width, avg_height))
        layout_list_t = []
        for i_f in range(ver_num):
            layout_t = QHBoxLayout()
            for i_ff in range(i_f*hor_num, min(obj_num_f, i_f*hor_num+hor_num)):
                name_f = self.df.loc[i_ff, 'Display_Name']
                exe_f = str(self.df.loc[i_ff, 'EXE_Path'])
                icon_f = self.df.loc[i_ff, 'Icon_Path'] if is_path(self.df.loc[i_ff, 'Icon_Path']) else r'E:\Super Launcher\icons\2k\button\opto_icon.svg'
                button = QPushButton()
                button.setIcon(QIcon(icon_f))  # 将图标文件路径替换为你实际的图标文件路径
                # 设置按钮样式
                button.setIcon(QIcon(icon_f))
                button.setIconSize(QSize(icon_size, icon_size))
                button.setFixedSize(button_size, button_size)
                button.setStyleSheet('''
                    QPushButton { 
                        border-radius: 20px; 
                        background-color: transparent; 
                        border: 0px solid #BFBFBF; 
                        } 
                    
                    QPushButton:hover { 
                        background-color: #619F99; 
                        }
                    
                    ''')
                button_name_f = f'{i_f}-{i_ff}-button'
                button.setObjectName(button_name_f)
                func_t = partial(self.launch_func, path_f=exe_f)
                self.click_effect[button_name_f] = func_t
                button.clicked.connect(self.button_flash)
                self.button_list[name_f] = button
                
                label = QLabel(name_f)
                label.setAlignment(Qt.AlignCenter)
                label.setFont(font_get(self.up.font_set['shortcut_button_title']))
                layout_tt = QVBoxLayout()
                layout_tt.addWidget(button)
                layout_tt.addWidget(label)
                layout_tt.setAlignment(Qt.AlignCenter)
                layout_t.addLayout(layout_tt)
                
            layout_list_t.insert(0, layout_t)
        for layout_i in layout_list_t:
            self.total_layout.addLayout(layout_i)
     # animation define function
    # animations define function
    def animations(self, button_f, size_ori, size_dst, func_f):
        self.animation1 = QPropertyAnimation(self, targetObject=button_f, propertyName=b'iconSize')
        self.animation1.setDuration(150)
        # self.animation1.setTargetObject(button_f)
        self.animation1.setStartValue(QSize(size_ori, size_ori))
        self.animation1.setEndValue(QSize(size_dst, size_dst))
        self.animation1.finished.connect(func_f)
        self.animation1.start()

    def launch_func(self, path_f):
        subprocess.Popen(['explorer', path_f])
        self.up.input_box.setFocus()
        # button click total func
    def button_flash(self):
        button_f = self.sender()
        func_f = self.click_effect[button_f.objectName()]
        size_ori = button_f.iconSize().width()
        size_dst = int(button_f.iconSize().width()*0.5)
        self.animations(button_f, size_ori, size_dst, lambda:self.animations(button_f, size_dst, size_ori, func_f))
# Shortcut Setting
class Shortcuts_setting(QWidget):
    def __init__(self, parent_f):
        super().__init__()
        self.up = parent_f
        self.setWindowFlags(self.windowFlags() | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setWindowTitle("Shortcut Setting")
        self.setWindowIcon(QIcon(self.up.path_set['Shortcut Setting']['window_button']))
        
        # self.setParent(parent_f)
        self.df = parent_f.shortcut_df
        
        self.col_num = 4
        self.num = len(parent_f.shortcut_df)
        
        self.icon_set_path = self.up.path_set['Shortcut Setting']['shortcut_button_set']
        self.coordinate_icon_set = self.up.path_set['Shortcut Setting']['coordinate_icon_set']
        self.default_icon_path = self.up.path_set['Shortcut Setting']['default_button_icon']
        self.exe_button_path = self.up.path_set['Shortcut Setting']['default_folder_icon']
        self.col_icon_path = {'coordinate':self.up.path_set['Shortcut Setting']['col_coordinate_icon'],
                            'name':self.up.path_set['Shortcut Setting']['col_name_icon'],
                            'icon':self.up.path_set['Shortcut Setting']['col_icon_icon'],
                            'path':self.up.path_set['Shortcut Setting']['col_path_icon'],
                            'search':self.up.path_set['Shortcut Setting']['col_search_icon']}
        
        self.data_load = {}
        self.coordinate_width = int(120*(self.up.size_set['res_x']))
        self.name_edit_width = int(270*(self.up.size_set['res_x']))
        self.pogram_icon_width = int(1.5*self.up.size_set['het'])
        self.exe_edit_width = int(700*(self.up.size_set['res_x']))
        self.exe_button_width = int(1.5*(self.up.size_set['het']))
        self.objs = {}
        self.effect_dict = {}
        
        x, y, w, h = self.up.get_geometry(self.up)
        x_1, y_1, w_1, h_1 = self.up.window_size_set_ori_data['Launcher']['main_window']
        self.w_1 = int(w_1*15/16)
        self.h_1 = int(h_1*16/15)
        self.setGeometry(x+w//2-w_1//2, y+h//2-h_1//2, self.w_1, self.h_1)
        
        self.layout_0 = QVBoxLayout()
        self.setLayout(self.layout_0)
        self.init_title()
        self.init_colname()
        self.init_obj()
        self.init_button()


        # For mouse control
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton:
            self.move(event.globalPos() - self.drag_position)
            event.accept()
    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            event.accept()
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        # 绘制带有圆角和渐变的窗口背景
        gradient = QLinearGradient(0, 0, 0, self.height())
        gradient.setColorAt(0, QColor(113, 148, 139))
        gradient.setColorAt(1, QColor(162, 245, 224))
        painter.setBrush(gradient)
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(self.rect(), 20, 20)
        
    
    def init_title(self):
        title_label = QLabel("Shortcut Settings")
        title_label.setFont(font_get(self.up.font_set['opto_setting_title']))
        title_label.setFixedHeight(int(1.5*self.up.size_set['het']))  # 指定标题高度
        title_label.setStyleSheet("background-color: transparent;")  # 设置标题透明背景
        title_label.setAlignment(Qt.AlignCenter)
        self.layout_0.addWidget(title_label)
    
    def init_colname(self):
        self.layout_col = QHBoxLayout()
        col_names_size = {'coordinate':self.coordinate_width, 
                          'name':self.name_edit_width, 
                          'icon':self.pogram_icon_width, 
                          'path':self.exe_edit_width, 
                          'search':self.exe_button_width}
        for name_i, width_i in col_names_size.items():
            col_label = QLabel()
            col_label.setPixmap(QIcon(self.col_icon_path[name_i]).pixmap(self.up.size_set['het'], self.up.size_set['het']))  # 设置图标的大小  
            col_label.setFixedSize(width_i, self.up.size_set['het'])  # 指定标题高度
            col_label.setStyleSheet("background-color: transparent;")  # 设置标题透明背景
            col_label.setAlignment(Qt.AlignCenter)
            self.layout_col.addWidget(col_label)
        layout_widget = QWidget()
        layout_widget.setLayout(self.layout_col)
        layout_widget.setFixedWidth(int(0.96*self.w_1))  # 设置布局的宽度为300像素
        self.layout_0.addWidget(layout_widget)
        
    def init_obj(self):
        obj_layout = QVBoxLayout()
        for i_f in range(self.num):
            self.data_load[i_f] = {'Display_Name':None,
                                   'Icon_Path':None,
                                   'EXE_Path':None}
            layout_row = QHBoxLayout()
            
            # coordinate define
            coordinate_name = f'{(i_f)%self.col_num+1}-{(i_f)//self.col_num+1}.svg'
            setattr(self, f'coordinate_label_{i_f}', QLabel())
            coordinate_label = getattr(self, f'coordinate_label_{i_f}')
            coordinate_label.setObjectName(f'coordinate_label_{i_f}')
            coor_icon_path = os.path.join(self.coordinate_icon_set[i_f])
            coordinate_label.setPixmap(QIcon(coor_icon_path).pixmap(self.up.size_set['het'], self.up.size_set['het']))  # 设置图标的大小  
            coordinate_label.setFixedSize(self.coordinate_width, self.up.size_set['het'])  # 指定标题高度
            coordinate_label.setStyleSheet('''background-color: transparent;
                                           border:none''')  # 设置标题透明背景
            coordinate_label.setAlignment(Qt.AlignCenter)
            # coordinate_label.setFixedSize( self.coordinate_width, int(1.2*self.up.size_set['het']))
            # coordinate_label.setFont(font_get(self.up.font_set['opto_setting_coordinate']))

            # coordinate_label.setStyleSheet("background-color: rgba(255, 255, 255, 45);border:none")
            layout_row.addWidget(coordinate_label)
            
            name_if = self.df.loc[i_f, 'Display_Name']
            name_if = name_if if isinstance(name_if, str) else 'Uncertain'
            setattr(self, f'name_edit_{i_f}', QLineEdit(name_if))
            name_edit = getattr(self, f'name_edit_{i_f}')
            name_edit.setObjectName(f'name_edit_{i_f}')
            name_edit.setFixedSize(self.name_edit_width, int(1.2*self.up.size_set['het']))
            name_edit.setFont(font_get(self.up.font_set['opto_setting_name_edit']))
            name_edit.setStyleSheet("background-color: transparent;")
            layout_row.addWidget(name_edit)
            self.data_load[i_f]['Display_Name'] = name_edit
            
            icon_if = self.df.loc[i_f, 'Icon_Path'] if is_path(self.df.loc[i_f, 'Icon_Path'], True) else self.default_icon_path
            setattr(self, f'icon_button_{i_f}', QPushButton())
            icon_button = getattr(self, f'icon_button_{i_f}')
            icon_button.setObjectName(f'icon_button_{i_f}')
            icon_button.setIcon(QIcon(icon_if))
            icon_button.setIconSize(QSize(self.up.size_set['het'], self.up.size_set['het']))
            icon_button.setFixedSize(self.pogram_icon_width, self.up.size_set['het'])
            self.data_load[i_f]['Icon_Path'] = icon_if
            icon_button.setStyleSheet('''QPushButton {
                                        background-color: transparent;
                                        border: none;
                                        padding: 10px 20px;
                                        }
    
                                        QPushButton:pressed {
                                            padding: 8px 18px;
                                        }
                                      ''')
            func_t = partial(self.icon_click, button_f=icon_button, row_f=i_f)
            self.effect_dict[f'icon_button_{i_f}'] = func_t
            icon_button.clicked.connect(self.button_flash)
            layout_row.addWidget(icon_button)
            
            exe_layout = QHBoxLayout()
            exe_f = self.df.loc[i_f, 'EXE_Path']
            exe_f = exe_f if isinstance(exe_f, str) else 'Uncertain.exe'
            
            setattr(self, f'exe_edit_{i_f}', QLineEdit())
            exe_edit = getattr(self, f'exe_edit_{i_f}')
            exe_edit.setObjectName(f'exe_edit_{i_f}')
            exe_edit.setFixedSize(self.exe_edit_width, int(1.2*self.up.size_set['het']))
            exe_edit.setPlaceholderText(exe_f)
            check_f = partial(self.check_path_validity, line_edit=exe_edit)
            exe_edit.textChanged.connect(check_f)
            exe_edit.setFont(font_get(self.up.font_set['opto_setting_exe_edit']))
            exe_edit.setStyleSheet("background-color: transparent;")
            layout_row.addWidget(exe_edit)
            self.data_load[i_f]['EXE_Path'] = exe_edit
            
            setattr(self, f'exe_button_{i_f}', QPushButton())
            exe_button = getattr(self, f'exe_button_{i_f}')
            exe_button.setObjectName(f'exe_button_{i_f}')
            exe_button.setIcon(QIcon(self.exe_button_path))
            exe_button.setIconSize(QSize(self.up.size_set['het'], self.up.size_set['het']))
            exe_button.setFixedSize(self.exe_button_width, int(1.2*self.up.size_set['het']))
            exe_button.setStyleSheet("background-color: transparent;border:none")
            func_t2 = partial(self.exe_search, line_f=exe_edit)
            self.effect_dict[f'exe_button_{i_f}'] = func_t2
            exe_button.clicked.connect(self.button_flash)
            layout_row.addWidget(exe_button)
            
            layout_row.addLayout(exe_layout)
            
            obj_layout.addLayout(layout_row)
        
        # Create a scroll area and add the content layout to it
        scroll_content = QWidget()  # Create a widget to contain the layout
        scroll_content.setLayout(obj_layout)
        
        frame = QFrame()  # Create a frame to contain the scroll area
        frame_layout = QVBoxLayout()
        frame_layout.addWidget(scroll_content)
        frame_layout.setContentsMargins(0, 0, 0, 0)
        frame.setLayout(frame_layout)
        frame.setStyleSheet("background: transparent;border-radius: 10px; border: 1px solid gray; background-color: transparent;")  # Set border to the frame
        
        scroll_area = QScrollArea()
        scroll_area.setWidget(frame)  # Set the frame as the widget of the scroll area
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("background: transparent; border: none;")  # Set background to transparent and no border to the scroll area
        self.layout_0.addWidget(scroll_area)
        
        # Customize scrollbar style
        scroll_bar = scroll_area.verticalScrollBar()
        scroll_bar.setStyleSheet("""
            QScrollBar:vertical {
                border: none;
                background: transparent;
                width: 20px;
                margin: 5px;
                
            }

            QScrollBar::handle:vertical {
                background: rgba(255, 255, 255, 25);
                min-height: 20px;
                border-radius: 5px;
            }

            QScrollBar::handle:vertical:hover {
                background: #737373;
            }

            QScrollBar::add-line:vertical {
                background: transparent;
                height: 0px;
                subcontrol-position: bottom;
                subcontrol-origin: margin;
            }

            QScrollBar::sub-line:vertical {
                background: transparent;
                height: 0px;
                subcontrol-position: top;
                subcontrol-origin: margin;
            }
            
            QScrollBar::sub-page:vertical, QScrollBar::add-page:vertical {
            background: transparent; /* 将滚动槽的背景色设置为透明 */
            }
        """)

    def init_button(self):
        button_layout = QHBoxLayout()
        # 创建确认和取消按钮
        confirm_button = QPushButton("Confirm")
        confirm_button.setFont(font_get({'Family':'Jetbrains Mono', 'PixelSize':30, 'Weight':70}))
        confirm_button.clicked.connect(self.confirm_action)
        confirm_button.setStyleSheet('''
        QPushButton {
            padding: 10px;
            color: #fff;
            border: none;
            border-radius: 10px;
            background-color: #286FB1;
        }

        QPushButton:hover {
            background-color: #2CC1E1; /* 设置悬停时的颜色 */
        }
        
        QPushButton:pressed {
                background-color: #349179;
            }
        
        ''')
        
        
        cancel_button = QPushButton("Cancel")
        cancel_button.setFont(font_get({'Family':'Jetbrains Mono', 'PixelSize':30, 'Weight':70}))
        cancel_button.clicked.connect(self.cancel_action)
        cancel_button.setStyleSheet('''
        QPushButton {
            padding: 10px;
            color: #fff;
            border: none;
            border-radius: 10px;
            background-color: #C60618;
        }

        QPushButton:hover {
            background-color: #FF4A6C; /* 设置悬停时的颜色 */
        }
        
        QPushButton:pressed {
            background-color: #349179;
        }
        
        ''')
        
        name_1 = "Launcher"
        launcher_excel_open_button = QPushButton(name_1)
        launcher_excel_open_button.setFont(font_get({'Family':'Jetbrains Mono', 'PixelSize':30, 'Weight':70}))
        func_1 = partial(self.open_func, sign_f=name_1)
        launcher_excel_open_button.clicked.connect(func_1)
        launcher_excel_open_button.setStyleSheet('''
        QPushButton {
            padding: 10px;
            color: #fff;
            border: none;
            border-radius: 10px;
            background-color: #44BBB1;
        }

        QPushButton:hover {
            background-color:#2CE1B3; /* 设置悬停时的颜色 */
        }
        
        QPushButton:pressed {
            background-color: #349179;
        }
        
        ''')
        
        name_2 = "Shortcut"
        shortcut_excel_open_button = QPushButton(name_2)
        shortcut_excel_open_button.setFont(font_get({'Family':'Jetbrains Mono', 'PixelSize':30, 'Weight':70}))
        func_2 = partial(self.open_func, sign_f=name_2)
        shortcut_excel_open_button.clicked.connect(func_2)
        shortcut_excel_open_button.setStyleSheet('''
            QPushButton {
                padding: 10px;
                color: #fff;
                border: none;
                border-radius: 10px;
                background-color: #44BBB1;
            }

            QPushButton:hover {
                background-color: #2CE1B3; /* 设置悬停时的颜色 */
            }
            QPushButton:pressed {
                background-color: #349179;
            }
            ''')
        
        button_layout.addWidget(launcher_excel_open_button)
        button_layout.addWidget(shortcut_excel_open_button)
        button_layout.addWidget(confirm_button)
        button_layout.addWidget(cancel_button)
        
        self.layout_0.addLayout(button_layout)
    
    def confirm_action(self):
        df_n = self.dict_trans()
        print(df_n)
        self.up.shortcut_df = df_n
        # relaunch button set
        self.up.shortcut_button.close()
        self.up.shortcut_button = Shortcuts(self.up)
        self.hide()
        # write in dataframe
        wb = openpyxl.load_workbook(self.up.path_set['Shortcut Setting']['shortcut_set'], data_only=True)
        ws = wb.worksheets[0]
        for row_i in range(len(df_n)):
            for col_i in range(3):
                ws.cell(row=row_i+2, column=col_i+1, value=df_n.iloc[row_i, col_i])
        wb.save(self.up.path_set['Shortcut Setting']['shortcut_set'])
        self.close()

    
    def dict_trans(self):
        df_copy = copy.deepcopy(self.df)
        for key_i, dict_i in self.data_load.items():
            for key_ii, value_i in dict_i.items():
                if key_ii in ['Display_Name', 'EXE_Path']:
                    if value_i.text():
                        df_copy.at[key_i, key_ii] = value_i.text()
                    else:
                        df_copy.at[key_i, key_ii] = self.df.loc[key_i, key_ii]
            for key_ii, value_i in dict_i.items():
                if key_ii == 'Icon_Path':
                    format_n = value_i.split('.')[-1]
                    obj_name = df_copy.loc[key_i, 'Display_Name']
                    path_new = os.path.join(os.path.dirname(self.icon_set_path[0]), f'{obj_name}.{format_n}')
                    if value_i != path_new:
                        shutil.copy(value_i, path_new)
                    df_copy.at[key_i, key_ii] = path_new
        return df_copy
    
    def cancel_action(self):
        self.close()
    
    def open_func(self, sign_f):
        if sign_f == 'Launcher':
            path_f = self.up.path_set['Launcher Mode']['launcher_set']
        elif sign_f == 'Shortcut':
            path_f = self.up.path_set['Shortcut Setting']['shortcut_set']
        else:
            path_f = ''
        subprocess.Popen(['explorer', path_f])
    # animation define function
    def animations(self, button_f, size_ori, size_dst, func_f):
        self.animation1 = QPropertyAnimation(self, targetObject=button_f, propertyName=b'iconSize')
        self.animation1.setDuration(120)
        # self.animation1.setTargetObject(button_f)
        self.animation1.setStartValue(QSize(size_ori, size_ori))
        self.animation1.setEndValue(QSize(size_dst, size_dst))
        self.animation1.finished.connect(func_f)
        self.animation1.start()
    # button click total func
    def button_flash(self):
        button_f = self.sender()
        func_f = self.effect_dict[button_f.objectName()]
        size_ori = button_f.iconSize().width()
        size_dst = int(button_f.iconSize().width()*0.5)
        self.animations(button_f, size_ori, size_dst, lambda:self.animations(button_f, size_dst, size_ori, func_f))
    # icon click conduct function
    def icon_click(self, button_f, row_f):
        filename_f, _ = QFileDialog.getOpenFileName(None, "Chose your Icon", "", "Photo File (*.png *.jpg *.bmp *.jepg *.svg *ico)")
        if not filename_f:
            return
        self.data_load[row_f]['Icon_Path'] = filename_f.replace('/', '\\')
        button_f.setIcon(QIcon(filename_f))
        button_f.setIconSize(QSize(60, 60))
    # check exe path validity
    def check_path_validity(self, text, line_edit):
        text_f = line_edit.text()
        if os.path.exists(text_f) or (not text_f):
            line_edit.setStyleSheet("border: 1px solid grey; background-color: transparent;")  # 路径不存在，变红
        elif os.path.isabs(text_f):
            line_edit.setStyleSheet("border: 3px solid #F0F411; background-color: transparent;")  # 路径不存在，变红
        else:
            line_edit.setStyleSheet("border: 3px solid #EE1515; background-color: transparent;")  # 路径存在，变黄
    # excel path search
    def exe_search(self, line_f):
        file_path_f, _ = QFileDialog.getOpenFileName(None, "Chose Your EXE Path", "", "All Files (*);;Folders")
        if file_path_f:
            line_f.setText(file_path_f.replace('/', '\\').strip())


# main funtion entry
def main(script_path_f):
    app = QApplication([])
    # main para set
    work_dir = os.getcwd()
    max_op_num = 99
    excel_path = r'.\load\Launcher Mode\launcher_set.xlsx'
    scrx, scry = get_screen_size()
    res_x, res_y = scrx/2560, scry/1440 # resize proportion is based on 2K resolution
    
    info_df = excel_to_df(excel_path, 'A:D', sheet_f='all')
    associate_class = Associate(program_info_df=info_df, ouput_num_f=max_op_num)

    size_set = {'screen':[scrx, scry],
                'res_x':res_x,
                'res_y':res_y,
                'main_window':[1600*res_x, 900*res_y],   # x,y
                'gap':int(50*res_x),   # common gap between parts or borders
                'het':int(60*res_y),   # height of input_box,switch_button and so on
                'srh_r':int(200*max(min({res_x, res_y}), 1)),# radius of search model Icon
    }
    
    num_set = {'associate_list':max_op_num,
               'horizontal_button_num':4}
    
    name_set = {'script_path':script_path_f,
                'window_name':['main_window',
                               'associate_list',
                               'gpt_box',
                               'input_box',
                               'switch_button',
                               'srh_icon',
                               'opto_box',
                               'opto_icon',
                               'shortcut_entry',
                               'shortcut_button',
                               'player',
                               ],
                'main_window':'Super Launcher',
                'mode_list':['Launcher Mode',
                              'Searcher Mode' ,
                              'Music Player',
                              'Opto System',
                              'Mistral-7B-Instruct-v0.2',
                              'CodeLlama-7b-Instruct-hf',
                              'Llama2-2-7b-chat-hf'
                              ],
                'mode_clarify':{'Launcher':['Launcher Mode'],
                                 'Searcher':['Searcher Mode'],
                                 'GPT':['Mistral-7B-Instruct-v0.2',
                                        'CodeLlama-7b-Instruct-hf',
                                        'Llama2-2-7b-chat-hf'],
                                 'Opto': ['Opto System'],
                                 'Media':['Music Player']}}
    
    func_set = {'associate_list':associate_class}
    
    font_set_ori = {'associate_list':{'Family':'微软雅黑', 'PixelSize':50},
                'input_box':{'Family':'微软雅黑', 'PixelSize':40},
                'gpt_box':{'Family':'微软雅黑', 'PixelSize':35},
                'opto_box':{'Family':'微软雅黑', 'PixelSize':35},
                'switch_button':{'Family':'Jetbrains Mono', 'PixelSize':50},
                'opto_setting_lineedit':{'Family':'微软雅黑', 'PixelSize':50},
                'opto_setting_title':{'Family':'Arial', 'PixelSize':80, 'Weight':80},
                'opto_setting_button':{'Family':'Jetbrains Mono', 'PixelSize':50, 'Weight':65},
                'shortcut_button_title':{'Family':'Jetbrains Mono', 'PixelSize':40, 'Weight':60},
                'shortcut_setting_button':{'Family':'Jetbrains Mono', 'PixelSize':60, 'Weight':70},
                'opto_setting_title':{'Family':'Arial', 'PixelSize':80, 'Weight':80},
                'opto_setting_colname':{'Family':'Arial', 'PixelSize':55, 'Weight':60},
                'opto_setting_coordinate':{'Family':'微软雅黑', 'PixelSize':55, 'Weight':60},
                'opto_setting_name_edit':{'Family':'Jetbrains Mono', 'PixelSize':60, 'Weight':70},
                'opto_setting_exe_edit':{'Family':'微软雅黑', 'PixelSize':45},
                }
    
    media_font_set = {'music_list_name':{'Family':'JetBrains Mono', 'PixelSize':42, 'Weight':70},
                      'search_box':{'Family':'微软雅黑', 'PixelSize':30},
                      'table_header':{'Family':'微软雅黑', 'PixelSize':30},
                      'number_label':{'Family':'微软雅黑', 'PixelSize':30},
                      'song_title':{'Family':'微软雅黑', 'PixelSize':30},
                      'artist_name':{'Family':'微软雅黑', 'PixelSize':26},
                      'duration_time':{'Family':'Consolas', 'PixelSize':28, 'Weight':60},
                      'file_format':{'Family':'Consolas', 'PixelSize':28, 'Weight':60},
                      'file_size':{'Family':'Consolas', 'PixelSize':28, 'Weight':60},
                      'song_title_bottom':{'Family':'微软雅黑', 'PixelSize':38, 'Weight':60},
                      'music_origin':{'Family':'JetBrains Mono', 'PixelSize':32},
                      'slider_label':{'Family':'Consolas', 'PixelSize':32},
                      'add_window_top_label':{'Family':'微软雅黑', 'PixelSize':30},
                      'add_window_list_name':{'Family':'微软雅黑', 'PixelSize':30},
                      'add_window_list_detail':{'Family':'微软雅黑', 'PixelSize':30},
                      'tip_label':{'Family':'微软雅黑', 'PixelSize':30},

                    }
    
    
    
    path_set = {
        'main_window':
                    {'taskbar_icon':None,
                     'maximum':None,
                     'minimum':None,
                     'close':None,
                     'middle':None,},
               
        'Launcher Mode':
                    {'launcher_set':None},
                
        'Shortcut Button':
                    dict(),
                
        'Shortcut Setting':
                    {'shortcut_set':None, 
                     'entry_button':None, 
                     'window_button':None,
                     'shortcut_button_set':'buttons',
                     'coordinate_icon_set':'coordinate',
                     'default_button_icon':None,
                     'default_folder_icon':None,
                     'col_coordinate_icon':None,
                     'col_name_icon':None,
                     'col_icon_icon':None,
                     'col_path_icon':None,
                     'col_search_icon':None,},
                
        'Opto Setting':
                    {'opto_set':None,
                     'entry_button':None},
                
        'Searcher Mode':
                    {'srh_icon_set':'srh_icons'},
                }
    media_path_set = {
        'default':{
                    'cover':None,
                    },
                
        'data':{'music_list_dict':None,
                'list_cover':None,
                'music_cover':None,
                'config':None,
                },
        
        'top':{'refresh':None,
               'search':None,
               },
        
        'mode':{'Favor':None,
                'History':None,
                'Local':None,
                'delete':None,
                'empty':None,
                'create':None,
                },
        
        'list':{'add':None,
                'delete':None,
                'like_grey':None,
                'like_red':None,
                'pause':None,
                'play':None,},
        
        'bottom':{'order':'order',
                'volume':'volume',
                'former':None,
                'next':None,
                'play':None,
                'suspend':None,},
            }   
    
    path_set = path_trans(path_set, os.path.join(work_dir, 'load'))
    media_path_set = path_trans(media_path_set, os.path.join(work_dir, 'load', 'player'))
    media_font_set = resize(media_font_set, size_set, 'font')
    window_size_set_ori = {
                    'Launcher':{
                            'main_window':['scrx/2-win_x/2','scry/2-win_y/2','win_x','win_y'],   # [x,y,w,h]
                            'switch_button':['gap','gap',0,'het'],
                            'srh_icon': None,
                            'input_box':['gap','1.5*gap+het','win_x-2*gap','het'],
                            'opto_icon': None,
                            'shortcut_entry': ['win_x-gap-het','1.5*gap+1.1*het','0.8*het','0.8*het'],
                            'shortcut_button':['(win_x+gap)/2', '1.8*gap+2*het','(win_x-3*gap)/2','win_y-2.8*gap-2*het'],
                            'associate_list':['gap','1.8*gap+2*het','win_x/2-gap','win_y-2.8*gap-2*het'],
                            'gpt_box':None,
                            'opto_box': None,
                            'opto_settings':None,},
                    
                    'Searcher':{
                            'main_window':['scrx/2-win_x/2','scry/2-win_y/2','win_x','win_y'],
                            'switch_button':['gap','gap',0,'het'],
                            'srh_icon':['win_x/2-radius/2','win_y/2-radius/2-3.6*gap','radius','radius'],
                            'input_box':['gap','win_y/2','win_x-2*gap','het'],
                            'opto_icon':None,
                            'shortcut_entry':None,
                            'shortcut_button':None,
                            'associate_list':None,
                            'gpt_box':None,
                            'opto_box': None,
                            'opto_settings':None,},
                    
                    'GPT':{
                            'main_window':['scrx/2-win_x/2','scry/2-win_y/2','win_x','win_y'],
                            'switch_button':['gap','gap',0,'het'],
                            'srh_icon': None,
                            'input_box':['gap','1.5*gap+het','win_x-2*gap','het'],
                            'opto_icon': None,
                            'shortcut_button':None,
                            'shortcut_entry':None,
                            'associate_list':None,
                            'gpt_box':['gap','1.8*gap+2*het','win_x-2*gap','win_y-2.8*gap-2*het'],
                            'opto_box': None,
                            'opto_settings':None,},
                    
                    'Opto':{
                            'main_window':['scrx/2-win_x/2','scry/2-win_y/2','win_x','win_y'],
                            'switch_button':['gap','gap',0,'het'],
                            'srh_icon': None,
                            'input_box':['gap','1.5*gap+het','win_x-2*gap','het'],
                            'opto_icon': ['win_x-gap-het','1.5*gap+1.1*het','0.8*het','0.8*het'],
                            'shortcut_button':None,
                            'shortcut_entry':None,
                            'associate_list':None,
                            'gpt_box':None,
                            'opto_box': ['gap','1.8*gap+2*het','win_x-2*gap','win_y-2.8*gap-2*het'],
                            'opto_settings': [0,0,'win_x*0.9','win_y*0.8'],},
                    
                    'Media':{
                            'main_window':['scrx/2-win_x*0.7','scry/2-win_y*0.7', '1.4*win_x','1.4*win_y'],
                            'switch_button':['gap','gap',0,'het'],
                            'player':['gap', '1.2*gap+het', 'win_x-2*gap', 'win_y-1.2*gap-het'],
                        },}
    # resize all geometry to fit in screen resolution
    # font_set = resize(font_set_ori, size_set, 'font')
    font_set = font_set_ori
    # QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    # QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    # QApplication.setAttribute(Qt.AA_UseSoftwareOpenGL)
    
    window_main = SuperLauncher(
                    num_set_f = num_set, 
                    name_set_f = name_set,
                    func_set_f = func_set,
                    font_set_f = font_set,
                    path_set_f = path_set,
                    workdir_f = work_dir,
                    window_size_set_ori_f = window_size_set_ori,
                    size_set_f = size_set, 
                    app_f = app,
                    media_path_set_f = media_path_set,
                    media_font_set_f = media_font_set)
    window_main.createTrayIcon()
    window_main.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    script_path = os.path.abspath(__file__) 
    main(script_path)